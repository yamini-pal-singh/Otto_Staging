#!/usr/bin/env python3
"""
Otto Intelligence — Gemini vs Production Comparison Test
=========================================================
Processes calls on STAGING (Gemini) and fetches existing PRODUCTION (OpenAI)
results, then compares all outputs field-by-field.

Generates:
  1. HTML dashboard — gemini_comparison_dashboard.html (stakeholder-friendly)
  2. Excel report   — gemini_comparison_report.xlsx
  3. JSON dump      — gemini_comparison_data.json (raw data for debugging)

Usage:
    python3 test_gemini_comparison.py
    python3 test_gemini_comparison.py --skip-submit   # only fetch & compare (calls already submitted)
    python3 test_gemini_comparison.py --batch-size 5   # submit 5 at a time
"""
import os
import sys
import uuid
import time
import json
import argparse
from datetime import datetime
from copy import copy

import requests

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl required: pip install openpyxl")
    sys.exit(1)

# ─────────────────────────────────────────────────────────────────────────────
# Configuration
# ─────────────────────────────────────────────────────────────────────────────
STAGING_BASE_URL = os.getenv("STAGING_BASE_URL", "https://ottoai-stage.shunyalabs.ai").rstrip("/")
PROD_BASE_URL = os.getenv("PROD_BASE_URL", "https://ottoai.shunyalabs.ai").rstrip("/")
API_KEY = os.getenv("OTTO_API_KEY", "5q3fwliU9ZFo3epTCsUfUiDw1Dy4DnBP")
COMPANY_ID = "1be5ea90-d3ae-4b03-8b05-f5679cd73bc4"

STAGING_HEADERS = {"X-API-Key": API_KEY, "Content-Type": "application/json"}
PROD_HEADERS = {"X-API-Key": API_KEY, "Content-Type": "application/json"}

TIMEOUT = 60
POLL_INTERVAL = 15  # seconds
MAX_POLL_TIME = 600  # 10 min per call

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_OUTPUT = os.path.join(OUTPUT_DIR, "gemini_comparison_report.xlsx")
JSON_OUTPUT = os.path.join(OUTPUT_DIR, "gemini_comparison_data.json")
HTML_OUTPUT = os.path.join(OUTPUT_DIR, "gemini_comparison_dashboard.html")

# ─────────────────────────────────────────────────────────────────────────────
# All 30 unique audio URLs from existing project sources
# Source 1: generate_new_report.py (10 URLs)
# Source 2: run_csr_test.py (10 URLs)
# Source 3: scripts/test_audio_urls.py (10 URLs)
# ─────────────────────────────────────────────────────────────────────────────
AUDIO_URLS = [
    # --- Source 1: generate_new_report.py ---
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/43ccc787-7f07-4a13-ac9d-d672d25a809c/4043504735.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/1fd7bea5-9ace-4e8f-a31f-152ea8269927/4015296617.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/bc946fa9-1e8a-4f1e-920c-64c4369fe778/4049722733.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/e2b0dab7-a05b-4448-8c13-fa3753f405ae/4036836500.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/8c6b15ee-5675-4e01-8b31-ff3658126353/4049722733.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/6e37c8bb-16bc-4e17-867e-ae5e9f57c3b9/4037028977.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/56dc7e30-ffed-4f8d-80eb-b514ffb30a50/4050591020.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/6566e3b9-acac-4b55-aad1-5742464107fa/4058579492.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/33cc8375-04ee-43a4-8644-bdffdb8d1b1b/4060192187.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/b9a3deca-d3e1-47ad-a2ad-875e58c0b7dc/4058565425.mp3",
    # --- Source 2: run_csr_test.py ---
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/bd24a254-fe4e-48ef-9c3e-3c3f3adfd68c/4082148782.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/00273c23-da2c-4e1b-b49c-4ff85d4a766a/4082112377.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/c0183543-60e4-434f-a247-89ec9ef8e1e3/4081871096.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/eb359c0a-02e2-4e94-b018-b3585c8a4024/4081765241.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/4da88a74-f129-481d-92d4-626c5a728835/4081742162.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/c6c4fff3-72d5-42d5-9cb4-16b8272e3d3b/4079556188.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/ddccb9c0-fce8-4eb5-81bd-5e738cb95979/4078581674.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/add722b1-c29f-4615-b389-4d29773d5958/st_309419607.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/df6c19b0-5d39-457b-8053-4dd83f4407cf/4076546354.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/6dd70da4-9d39-4496-871e-827af1429e29/4077402764.mp3",
    # --- Source 3: scripts/test_audio_urls.py ---
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/0caf5493-4ea1-4298-bee3-5ea9a83a2431/4094726633.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/31047617-96e2-4b71-8167-9b6aa3acaa3e/4094713955.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/cd5faebf-2c6b-4733-b560-e2a97eec644f/4094604863.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/90684b34-3972-4abe-a56b-07633b75da39/4094588411.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/349c746c-0a9b-44ff-a6c3-8c052a1efd29/4094548208.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/bdd04264-1591-4b4a-8aca-6f63beb79aba/4094484647.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/c979ee91-9189-4473-bcae-898ddd5d0028/4094457467.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/203c3c8b-4cce-4dfe-8ef7-74c4dd9ae3db/4094238074.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/1a6fa204-73b1-42e6-9adc-930a8be7196c/4093544600.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/cd8c80e2-0388-4ac7-818c-6f543bb500e8/4094042249.mp3",
]


# ─────────────────────────────────────────────────────────────────────────────
# API Helpers
# ─────────────────────────────────────────────────────────────────────────────

def health_check(base_url, headers, label):
    """Check if environment is reachable."""
    try:
        r = requests.get(f"{base_url}/health", headers=headers, timeout=10)
        ok = r.status_code == 200
        print(f"  {label}: HTTP {r.status_code} {'OK' if ok else 'FAIL'}")
        return ok
    except Exception as e:
        print(f"  {label}: UNREACHABLE ({e})")
        return False


def submit_call(base_url, headers, audio_url, idx):
    """Submit a call for processing. Returns (call_id, job_id) or (call_id, None)."""
    call_id = f"gemini_cmp_{idx+1}_{uuid.uuid4().hex[:8]}"
    payload = {
        "call_id": call_id,
        "company_id": COMPANY_ID,
        "audio_url": audio_url,
        "phone_number": "+14805551234",
        "rep_role": "customer_rep",
        "allow_reprocess": True,
        "metadata": {
            "agent": {
                "id": "USR_ANTHONY_ARIZONA",
                "name": "Anthony",
                "email": "anthony@arizonaroofers.com",
            }
        },
    }
    try:
        r = requests.post(
            f"{base_url}/api/v1/call-processing/process",
            headers=headers, json=payload, timeout=TIMEOUT,
        )
        data = r.json()
        job_id = data.get("job_id")
        print(f"  [{idx+1:2d}] HTTP {r.status_code} | call_id={call_id} | job_id={job_id}")
        if r.status_code in (200, 202):
            return call_id, job_id
        elif r.status_code == 409:
            print(f"       409 Conflict: {str(data.get('detail', ''))[:100]}")
            return call_id, job_id
        else:
            print(f"       Error: {str(data.get('detail', r.text[:200]))}")
            return call_id, None
    except Exception as e:
        print(f"  [{idx+1:2d}] Exception: {e}")
        return call_id, None


def poll_job(base_url, headers, job_id, label=""):
    """Poll job status until completed/failed or timeout. Returns (status, duration_sec)."""
    if not job_id:
        return "no_job_id", 0
    start = time.time()
    while time.time() - start < MAX_POLL_TIME:
        try:
            r = requests.get(
                f"{base_url}/api/v1/call-processing/status/{job_id}",
                headers=headers, timeout=TIMEOUT,
            )
            data = r.json()
            status = data.get("status", "unknown")
            progress = data.get("progress", {})
            pct = progress.get("percent", "?")
            step = progress.get("current_step", "?")
            print(f"    {label} status={status} progress={pct}% step={step}")
            if status in ("completed", "failed"):
                return status, round(time.time() - start, 1)
        except Exception as e:
            print(f"    {label} poll error: {e}")
        time.sleep(POLL_INTERVAL)
    return "timeout", round(time.time() - start, 1)


def fetch_summary(base_url, headers, call_id):
    """Fetch full summary for a call."""
    try:
        r = requests.get(
            f"{base_url}/api/v1/call-processing/summary/{call_id}",
            headers=headers, params={"include_chunks": "true"}, timeout=TIMEOUT,
        )
        if r.status_code == 200:
            return r.json()
        print(f"    Summary fetch HTTP {r.status_code} for {call_id}")
    except Exception as e:
        print(f"    Summary fetch error for {call_id}: {e}")
    return None


def fetch_detail(base_url, headers, call_id):
    """Fetch call detail with transcript and segments."""
    try:
        r = requests.get(
            f"{base_url}/api/v1/call-processing/calls/{call_id}/detail",
            headers=headers,
            params={"include_transcript": "true", "include_segments": "true"},
            timeout=TIMEOUT,
        )
        if r.status_code == 200:
            return r.json()
        print(f"    Detail fetch HTTP {r.status_code} for {call_id}")
    except Exception as e:
        print(f"    Detail fetch error for {call_id}: {e}")
    return None


def find_prod_call_by_audio(audio_url):
    """Search production for a completed call matching this audio URL."""
    try:
        # Search across multiple pages
        for offset in range(0, 500, 50):
            r = requests.get(
                f"{PROD_BASE_URL}/api/v1/call-processing/calls",
                headers=PROD_HEADERS,
                params={
                    "company_id": COMPANY_ID,
                    "limit": 50,
                    "offset": offset,
                    "sort_by": "call_date",
                    "sort_order": "desc",
                },
                timeout=TIMEOUT,
            )
            if r.status_code != 200:
                break
            calls = r.json().get("calls", [])
            if not calls:
                break
            for c in calls:
                if c.get("audio_url") == audio_url and c.get("status") == "completed":
                    return c.get("call_id")
    except Exception as e:
        print(f"    Prod search error: {e}")
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Data Extraction — Flat row from API response
# ─────────────────────────────────────────────────────────────────────────────

def safe_get(d, *keys, default=""):
    """Safely navigate nested dicts."""
    current = d
    for k in keys:
        if isinstance(current, dict):
            current = current.get(k)
        else:
            return default
        if current is None:
            return default
    return current


def extract_row(summary, detail, prefix=""):
    """Extract a flat dict of all fields from summary + detail responses.
    prefix: 'PROD_' or 'STAGING_' to namespace columns.
    """
    row = {}
    p = prefix

    if not summary and not detail:
        row[f"{p}Status"] = "NO DATA"
        return row

    row[f"{p}Status"] = "OK"

    # ── Summary ──
    s = safe_get(summary, "summary", default={})
    if isinstance(s, dict):
        row[f"{p}Summary"] = s.get("summary", "")
        row[f"{p}Key Points"] = "\n".join(s.get("key_points", []) or [])
        row[f"{p}Action Items"] = "\n".join(s.get("action_items", []) or [])
        row[f"{p}Next Steps"] = "\n".join(s.get("next_steps", []) or [])
        row[f"{p}Sentiment Score"] = s.get("sentiment_score")
        row[f"{p}Confidence Score"] = s.get("confidence_score")
        pending = s.get("pending_actions", []) or []
        row[f"{p}Pending Actions Count"] = len(pending)
        if pending:
            pa_texts = []
            for pa in pending:
                pa_texts.append(f"[{pa.get('type','')}] {pa.get('action_item','')} (owner: {pa.get('owner','')}, due: {pa.get('due_at','')})")
            row[f"{p}Pending Actions"] = "\n".join(pa_texts)

    # ── Qualification ──
    qual = safe_get(summary, "qualification", default={})
    if isinstance(qual, dict):
        row[f"{p}Customer Name"] = qual.get("customer_name", "")
        row[f"{p}Customer Phone"] = qual.get("customer_phone", "")
        row[f"{p}Customer Email"] = qual.get("customer_email", "")
        row[f"{p}Qualification Status"] = qual.get("qualification_status", "")
        row[f"{p}Service Requested"] = qual.get("service_requested", "")
        row[f"{p}Booking Status"] = qual.get("booking_status", "")
        row[f"{p}Call Type"] = qual.get("detected_call_type", "")
        row[f"{p}Customer Type"] = "Existing" if qual.get("is_existing_customer") else "New" if qual.get("is_existing_customer") is False else ""
        row[f"{p}Call Outcome"] = qual.get("call_outcome_category", "")
        row[f"{p}Appointment Confirmed"] = qual.get("appointment_confirmed")
        row[f"{p}Appointment Date"] = qual.get("appointment_date", "")
        row[f"{p}Appointment Type"] = qual.get("appointment_type", "")
        row[f"{p}Follow Up Required"] = qual.get("follow_up_required")
        row[f"{p}Follow Up Reason"] = qual.get("follow_up_reason", "")
        row[f"{p}Confidence Score (Qual)"] = qual.get("confidence_score")

        # Scope
        row[f"{p}Scope Classification"] = qual.get("scope_classification", "")
        row[f"{p}Scope Reason"] = qual.get("scope_reason", "")
        row[f"{p}Scope Confidence"] = qual.get("scope_confidence")

        # BANT
        bant = qual.get("bant_scores", {}) or {}
        row[f"{p}BANT Budget"] = bant.get("budget")
        row[f"{p}BANT Authority"] = bant.get("authority")
        row[f"{p}BANT Need"] = bant.get("need")
        row[f"{p}BANT Timeline"] = bant.get("timeline")
        row[f"{p}BANT Overall"] = qual.get("overall_score")

        # Address
        row[f"{p}Service Address Raw"] = qual.get("service_address_raw", "")
        addr = qual.get("service_address_structured", {}) or {}
        if addr:
            row[f"{p}Address City"] = addr.get("city", "")
            row[f"{p}Address State"] = addr.get("state", "")

        # Property
        prop = qual.get("property_details", {}) or {}
        if prop:
            row[f"{p}Roof Type"] = prop.get("roof_type", "")
            row[f"{p}HOA Status"] = prop.get("hoa_status", "")
            row[f"{p}Has Solar"] = prop.get("has_solar")

    # ── Lead Score ──
    lead = safe_get(summary, "lead_score", default={})
    if isinstance(lead, dict):
        row[f"{p}Lead Score"] = lead.get("total_score")
        row[f"{p}Lead Band"] = lead.get("lead_band", "")

    # ── Compliance ──
    comp = safe_get(summary, "compliance", default={})
    if isinstance(comp, dict):
        sop = comp.get("sop_compliance", {}) or {}
        row[f"{p}Compliance Score"] = sop.get("score")
        row[f"{p}Compliance Rate"] = sop.get("compliance_rate")
        stages = sop.get("stages", {}) or {}
        row[f"{p}Stages Total"] = stages.get("total")
        row[f"{p}Stages Followed"] = ", ".join(stages.get("followed", []) or [])
        row[f"{p}Stages Missed"] = ", ".join(stages.get("missed", []) or [])
        row[f"{p}Compliance Issues"] = "\n".join(sop.get("issues", []) or [])
        row[f"{p}Positive Behaviors"] = "\n".join(sop.get("positive_behaviors", []) or [])

        coaching_issues = sop.get("coaching_issues", []) or []
        if coaching_issues:
            ci_texts = []
            for ci in coaching_issues:
                ci_texts.append(f"[{ci.get('severity','')}] {ci.get('issue','')} | Fix: {ci.get('how_to_fix','')}")
            row[f"{p}Coaching Issues"] = "\n".join(ci_texts)

        coaching_strengths = sop.get("coaching_strengths", []) or []
        if coaching_strengths:
            cs_texts = []
            for cs in coaching_strengths:
                cs_texts.append(f"{cs.get('behavior','')} | {cs.get('why_effective','')}")
            row[f"{p}Coaching Strengths"] = "\n".join(cs_texts)

    # ── Objections ──
    obj_data = safe_get(summary, "objections", default={})
    if isinstance(obj_data, dict):
        objections = obj_data.get("objections", []) or []
        row[f"{p}Objections Count"] = obj_data.get("total_count", len(objections))
        if objections:
            obj_texts = []
            for o in objections:
                overcome = "Resolved" if o.get("overcome") else "Unresolved"
                obj_texts.append(
                    f"[{o.get('category_text','')}] \"{o.get('objection_text','')}\" "
                    f"({overcome}, severity={o.get('severity','')})"
                )
            row[f"{p}Objections Detail"] = "\n".join(obj_texts)

    # ── Transcript ──
    if detail:
        transcript = detail.get("transcript", "")
        if isinstance(transcript, dict):
            segments = transcript.get("segments", []) or []
            seg_texts = [f"[{seg.get('speaker','?')}] {seg.get('text','')}" for seg in segments]
            row[f"{p}Transcript"] = "\n".join(seg_texts)
            row[f"{p}Segment Count"] = len(segments)
        elif isinstance(transcript, str):
            row[f"{p}Transcript"] = transcript
            row[f"{p}Segment Count"] = len(detail.get("segments", []) or [])
        else:
            segments = detail.get("segments", []) or []
            seg_texts = [f"[{seg.get('speaker','?')}] {seg.get('text','')}" for seg in segments]
            row[f"{p}Transcript"] = "\n".join(seg_texts)
            row[f"{p}Segment Count"] = len(segments)

        # Check speaker labels
        segments = detail.get("segments", []) or []
        if not segments:
            t = detail.get("transcript", {})
            if isinstance(t, dict):
                segments = t.get("segments", []) or []
        speakers = set(seg.get("speaker", "") for seg in segments)
        row[f"{p}Speaker Labels"] = ", ".join(sorted(speakers))
        row[f"{p}Transcript Length"] = len(row.get(f"{p}Transcript", ""))

    return row


# ─────────────────────────────────────────────────────────────────────────────
# Comparison Logic
# ─────────────────────────────────────────────────────────────────────────────

def compare_fields(prod_row, staging_row):
    """Compare production vs staging fields and generate issues list."""
    issues = []

    def delta(field, threshold, severity="Medium"):
        pv = prod_row.get(f"PROD_{field}")
        sv = staging_row.get(f"STAGING_{field}")
        if pv is not None and sv is not None:
            try:
                d = abs(float(sv) - float(pv))
                if d > threshold:
                    issues.append({
                        "field": field,
                        "severity": severity,
                        "type": "Score Delta",
                        "detail": f"PROD={pv} STAGING={sv} delta={d:.3f} (threshold={threshold})",
                    })
                return d
            except (ValueError, TypeError):
                pass
        return None

    def match(field, severity="Medium"):
        pv = str(prod_row.get(f"PROD_{field}", "")).strip().lower()
        sv = str(staging_row.get(f"STAGING_{field}", "")).strip().lower()
        if pv and sv and pv != sv:
            issues.append({
                "field": field,
                "severity": severity,
                "type": "Mismatch",
                "detail": f"PROD='{pv}' STAGING='{sv}'",
            })
            return False
        return True

    def present(field, severity="High"):
        sv = staging_row.get(f"STAGING_{field}")
        pv = prod_row.get(f"PROD_{field}")
        if pv and not sv:
            issues.append({
                "field": field,
                "severity": severity,
                "type": "Missing in Staging",
                "detail": f"Present in PROD but missing/empty in STAGING",
            })
            return False
        return True

    # ── Score deltas ──
    delta("Compliance Score", 0.15, "High")
    delta("Sentiment Score", 0.3, "Medium")
    delta("Confidence Score", 0.2, "Medium")
    delta("BANT Budget", 0.2, "Medium")
    delta("BANT Authority", 0.2, "Medium")
    delta("BANT Need", 0.2, "Medium")
    delta("BANT Timeline", 0.2, "Medium")
    delta("BANT Overall", 0.2, "Medium")
    delta("Lead Score", 15, "Medium")
    delta("Compliance Rate", 0.15, "Medium")

    # ── Exact matches ──
    match("Booking Status", "High")
    match("Call Type", "High")
    match("Qualification Status", "Medium")
    match("Customer Type", "Low")
    match("Call Outcome", "Medium")
    match("Scope Classification", "Medium")

    # ── Field presence ──
    present("Summary", "High")
    present("Key Points", "High")
    present("Compliance Score", "High")
    present("Objections Count", "Medium")
    present("Transcript", "High")

    # ── Objection count delta ──
    pc = prod_row.get("PROD_Objections Count")
    sc = staging_row.get("STAGING_Objections Count")
    if pc is not None and sc is not None:
        try:
            od = abs(int(sc) - int(pc))
            if od > 2:
                issues.append({
                    "field": "Objections Count",
                    "severity": "Medium",
                    "type": "Count Delta",
                    "detail": f"PROD={pc} STAGING={sc} delta={od}",
                })
        except (ValueError, TypeError):
            pass

    # ── Staging errors ──
    if staging_row.get("STAGING_Status") == "NO DATA":
        issues.append({
            "field": "Processing",
            "severity": "High",
            "type": "Processing Failed",
            "detail": "Staging returned no data for this call",
        })

    return issues


# ─────────────────────────────────────────────────────────────────────────────
# Excel Report Writer
# ─────────────────────────────────────────────────────────────────────────────

# Styles
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
HEADER_FILL_META = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FILL_PROD = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
HEADER_FILL_STAGING = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
HEADER_FILL_COMPARE = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGN = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

SEVERITY_FILLS = {
    "High": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "Medium": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "Low": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
}

# Column definitions: (header, key, group)
COLUMNS = [
    # Meta
    ("Call #", "call_num", "meta"),
    ("Audio URL", "audio_url", "meta"),
    ("Source Batch", "source_batch", "meta"),
    ("PROD Call ID", "prod_call_id", "meta"),
    ("STAGING Call ID", "staging_call_id", "meta"),
    ("Staging Process Status", "staging_process_status", "meta"),
    ("Staging Process Time (s)", "staging_process_time", "meta"),
    # Production fields
    ("PROD Summary", "PROD_Summary", "prod"),
    ("PROD Key Points", "PROD_Key Points", "prod"),
    ("PROD Action Items", "PROD_Action Items", "prod"),
    ("PROD Next Steps", "PROD_Next Steps", "prod"),
    ("PROD Sentiment Score", "PROD_Sentiment Score", "prod"),
    ("PROD Confidence Score", "PROD_Confidence Score", "prod"),
    ("PROD Pending Actions Count", "PROD_Pending Actions Count", "prod"),
    ("PROD Pending Actions", "PROD_Pending Actions", "prod"),
    ("PROD Customer Name", "PROD_Customer Name", "prod"),
    ("PROD Customer Phone", "PROD_Customer Phone", "prod"),
    ("PROD Qualification Status", "PROD_Qualification Status", "prod"),
    ("PROD Service Requested", "PROD_Service Requested", "prod"),
    ("PROD Booking Status", "PROD_Booking Status", "prod"),
    ("PROD Call Type", "PROD_Call Type", "prod"),
    ("PROD Customer Type", "PROD_Customer Type", "prod"),
    ("PROD Call Outcome", "PROD_Call Outcome", "prod"),
    ("PROD Appointment Confirmed", "PROD_Appointment Confirmed", "prod"),
    ("PROD Appointment Date", "PROD_Appointment Date", "prod"),
    ("PROD Follow Up Required", "PROD_Follow Up Required", "prod"),
    ("PROD Scope Classification", "PROD_Scope Classification", "prod"),
    ("PROD BANT Budget", "PROD_BANT Budget", "prod"),
    ("PROD BANT Authority", "PROD_BANT Authority", "prod"),
    ("PROD BANT Need", "PROD_BANT Need", "prod"),
    ("PROD BANT Timeline", "PROD_BANT Timeline", "prod"),
    ("PROD BANT Overall", "PROD_BANT Overall", "prod"),
    ("PROD Lead Score", "PROD_Lead Score", "prod"),
    ("PROD Lead Band", "PROD_Lead Band", "prod"),
    ("PROD Compliance Score", "PROD_Compliance Score", "prod"),
    ("PROD Compliance Rate", "PROD_Compliance Rate", "prod"),
    ("PROD Stages Followed", "PROD_Stages Followed", "prod"),
    ("PROD Stages Missed", "PROD_Stages Missed", "prod"),
    ("PROD Coaching Issues", "PROD_Coaching Issues", "prod"),
    ("PROD Objections Count", "PROD_Objections Count", "prod"),
    ("PROD Objections Detail", "PROD_Objections Detail", "prod"),
    ("PROD Transcript Length", "PROD_Transcript Length", "prod"),
    ("PROD Segment Count", "PROD_Segment Count", "prod"),
    ("PROD Speaker Labels", "PROD_Speaker Labels", "prod"),
    # Staging fields (same order)
    ("STAGING Summary", "STAGING_Summary", "staging"),
    ("STAGING Key Points", "STAGING_Key Points", "staging"),
    ("STAGING Action Items", "STAGING_Action Items", "staging"),
    ("STAGING Next Steps", "STAGING_Next Steps", "staging"),
    ("STAGING Sentiment Score", "STAGING_Sentiment Score", "staging"),
    ("STAGING Confidence Score", "STAGING_Confidence Score", "staging"),
    ("STAGING Pending Actions Count", "STAGING_Pending Actions Count", "staging"),
    ("STAGING Pending Actions", "STAGING_Pending Actions", "staging"),
    ("STAGING Customer Name", "STAGING_Customer Name", "staging"),
    ("STAGING Customer Phone", "STAGING_Customer Phone", "staging"),
    ("STAGING Qualification Status", "STAGING_Qualification Status", "staging"),
    ("STAGING Service Requested", "STAGING_Service Requested", "staging"),
    ("STAGING Booking Status", "STAGING_Booking Status", "staging"),
    ("STAGING Call Type", "STAGING_Call Type", "staging"),
    ("STAGING Customer Type", "STAGING_Customer Type", "staging"),
    ("STAGING Call Outcome", "STAGING_Call Outcome", "staging"),
    ("STAGING Appointment Confirmed", "STAGING_Appointment Confirmed", "staging"),
    ("STAGING Appointment Date", "STAGING_Appointment Date", "staging"),
    ("STAGING Follow Up Required", "STAGING_Follow Up Required", "staging"),
    ("STAGING Scope Classification", "STAGING_Scope Classification", "staging"),
    ("STAGING BANT Budget", "STAGING_BANT Budget", "staging"),
    ("STAGING BANT Authority", "STAGING_BANT Authority", "staging"),
    ("STAGING BANT Need", "STAGING_BANT Need", "staging"),
    ("STAGING BANT Timeline", "STAGING_BANT Timeline", "staging"),
    ("STAGING BANT Overall", "STAGING_BANT Overall", "staging"),
    ("STAGING Lead Score", "STAGING_Lead Score", "staging"),
    ("STAGING Lead Band", "STAGING_Lead Band", "staging"),
    ("STAGING Compliance Score", "STAGING_Compliance Score", "staging"),
    ("STAGING Compliance Rate", "STAGING_Compliance Rate", "staging"),
    ("STAGING Stages Followed", "STAGING_Stages Followed", "staging"),
    ("STAGING Stages Missed", "STAGING_Stages Missed", "staging"),
    ("STAGING Coaching Issues", "STAGING_Coaching Issues", "staging"),
    ("STAGING Objections Count", "STAGING_Objections Count", "staging"),
    ("STAGING Objections Detail", "STAGING_Objections Detail", "staging"),
    ("STAGING Transcript Length", "STAGING_Transcript Length", "staging"),
    ("STAGING Segment Count", "STAGING_Segment Count", "staging"),
    ("STAGING Speaker Labels", "STAGING_Speaker Labels", "staging"),
    # Comparison
    ("Issues Count", "issues_count", "compare"),
    ("High Severity Issues", "high_issues", "compare"),
    ("Medium Severity Issues", "medium_issues", "compare"),
    ("Low Severity Issues", "low_issues", "compare"),
    ("Issues Detail", "issues_detail", "compare"),
    ("Compliance Score Delta", "delta_compliance", "compare"),
    ("Sentiment Score Delta", "delta_sentiment", "compare"),
    ("BANT Overall Delta", "delta_bant_overall", "compare"),
    ("Lead Score Delta", "delta_lead_score", "compare"),
    ("Booking Status Match", "match_booking", "compare"),
    ("Call Type Match", "match_call_type", "compare"),
    ("Overall Verdict", "verdict", "compare"),
]

GROUP_FILLS = {
    "meta": HEADER_FILL_META,
    "prod": HEADER_FILL_PROD,
    "staging": HEADER_FILL_STAGING,
    "compare": HEADER_FILL_COMPARE,
}


def safe_delta(prod_row, staging_row, field):
    """Calculate delta between prod and staging for a numeric field."""
    pv = prod_row.get(f"PROD_{field}")
    sv = staging_row.get(f"STAGING_{field}")
    if pv is not None and sv is not None:
        try:
            return round(abs(float(sv) - float(pv)), 4)
        except (ValueError, TypeError):
            pass
    return ""


def safe_match(prod_row, staging_row, field):
    """Check if prod and staging values match for a field."""
    pv = str(prod_row.get(f"PROD_{field}", "")).strip().lower()
    sv = str(staging_row.get(f"STAGING_{field}", "")).strip().lower()
    if not pv and not sv:
        return "N/A"
    if pv == sv:
        return "MATCH"
    return "MISMATCH"


def build_comparison_row(idx, audio_url, prod_call_id, staging_call_id,
                         staging_status, staging_time,
                         prod_row, staging_row, issues):
    """Build a single flat row dict for the Excel sheet."""
    row = {}
    row["call_num"] = idx + 1
    row["audio_url"] = audio_url

    # Determine source batch
    if idx < 10:
        row["source_batch"] = "Batch 1 (generate_new_report)"
    elif idx < 20:
        row["source_batch"] = "Batch 2 (run_csr_test)"
    else:
        row["source_batch"] = "Batch 3 (test_audio_urls)"

    row["prod_call_id"] = prod_call_id or "NOT FOUND"
    row["staging_call_id"] = staging_call_id or "NOT SUBMITTED"
    row["staging_process_status"] = staging_status or ""
    row["staging_process_time"] = staging_time or ""

    # Merge prod and staging extracted fields
    row.update(prod_row)
    row.update(staging_row)

    # Comparison fields
    high = [i for i in issues if i["severity"] == "High"]
    med = [i for i in issues if i["severity"] == "Medium"]
    low = [i for i in issues if i["severity"] == "Low"]

    row["issues_count"] = len(issues)
    row["high_issues"] = len(high)
    row["medium_issues"] = len(med)
    row["low_issues"] = len(low)
    row["issues_detail"] = "\n".join(
        f"[{i['severity']}] {i['field']}: {i['type']} — {i['detail']}" for i in issues
    )

    row["delta_compliance"] = safe_delta(prod_row, staging_row, "Compliance Score")
    row["delta_sentiment"] = safe_delta(prod_row, staging_row, "Sentiment Score")
    row["delta_bant_overall"] = safe_delta(prod_row, staging_row, "BANT Overall")
    row["delta_lead_score"] = safe_delta(prod_row, staging_row, "Lead Score")
    row["match_booking"] = safe_match(prod_row, staging_row, "Booking Status")
    row["match_call_type"] = safe_match(prod_row, staging_row, "Call Type")

    # Overall verdict
    if len(high) > 0:
        row["verdict"] = "FAIL"
    elif len(med) > 2:
        row["verdict"] = "WARN"
    else:
        row["verdict"] = "PASS"

    return row


def write_excel(all_rows, output_path):
    """Write the comparison report to Excel with two sheets:
    1. Per-Call Comparison — one row per call with all fields
    2. Executive Summary — aggregated stats
    """
    wb = openpyxl.Workbook()

    # ══════════════════════════════════════════════════════════════════════
    # Sheet 1: Per-Call Comparison
    # ══════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Per-Call Comparison"
    ws.freeze_panes = "A2"

    # Write headers
    for col_idx, (header, key, group) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = GROUP_FILLS[group]
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # Write data rows
    for row_idx, row_data in enumerate(all_rows, 2):
        for col_idx, (header, key, group) in enumerate(COLUMNS, 1):
            val = row_data.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = CELL_ALIGN
            cell.border = THIN_BORDER

        # Color-code verdict column
        verdict_col = next(i for i, (h, k, g) in enumerate(COLUMNS, 1) if k == "verdict")
        verdict_val = row_data.get("verdict", "")
        verdict_cell = ws.cell(row=row_idx, column=verdict_col)
        if verdict_val == "FAIL":
            verdict_cell.fill = SEVERITY_FILLS["High"]
            verdict_cell.font = Font(bold=True, color="9C0006")
        elif verdict_val == "WARN":
            verdict_cell.fill = SEVERITY_FILLS["Medium"]
            verdict_cell.font = Font(bold=True, color="9C6500")
        elif verdict_val == "PASS":
            verdict_cell.fill = SEVERITY_FILLS["Low"]
            verdict_cell.font = Font(bold=True, color="006100")

        # Color-code mismatch cells
        for col_idx, (header, key, group) in enumerate(COLUMNS, 1):
            if key.startswith("match_"):
                val = row_data.get(key, "")
                cell = ws.cell(row=row_idx, column=col_idx)
                if val == "MISMATCH":
                    cell.fill = SEVERITY_FILLS["High"]
                elif val == "MATCH":
                    cell.fill = SEVERITY_FILLS["Low"]

    # Auto-width (capped)
    for col_idx in range(1, len(COLUMNS) + 1):
        max_len = len(str(ws.cell(1, col_idx).value or ""))
        for row_idx in range(2, min(len(all_rows) + 2, 10)):
            val = str(ws.cell(row_idx, col_idx).value or "")
            max_len = max(max_len, min(len(val), 50))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 55)

    # ══════════════════════════════════════════════════════════════════════
    # Sheet 2: Executive Summary
    # ══════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Executive Summary")

    total = len(all_rows)
    passed = sum(1 for r in all_rows if r.get("verdict") == "PASS")
    warned = sum(1 for r in all_rows if r.get("verdict") == "WARN")
    failed = sum(1 for r in all_rows if r.get("verdict") == "FAIL")
    staging_completed = sum(1 for r in all_rows if r.get("staging_process_status") == "completed")
    prod_found = sum(1 for r in all_rows if r.get("prod_call_id", "NOT FOUND") != "NOT FOUND")

    total_high = sum(r.get("high_issues", 0) for r in all_rows)
    total_med = sum(r.get("medium_issues", 0) for r in all_rows)
    total_low = sum(r.get("low_issues", 0) for r in all_rows)

    # Averages for numeric fields
    def avg_field(rows, key):
        vals = [r.get(key) for r in rows if r.get(key) is not None and r.get(key) != ""]
        nums = []
        for v in vals:
            try:
                nums.append(float(v))
            except (ValueError, TypeError):
                pass
        return round(sum(nums) / len(nums), 4) if nums else "N/A"

    summary_data = [
        ("GEMINI MIGRATION — EXECUTIVE SUMMARY", "", ""),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), ""),
        ("", "", ""),
        ("METRIC", "VALUE", "STATUS"),
        ("Total Calls Tested", total, ""),
        ("Production Baseline Found", f"{prod_found}/{total}", "PASS" if prod_found == total else "WARN"),
        ("Staging Processed Successfully", f"{staging_completed}/{total}", "PASS" if staging_completed == total else "FAIL"),
        ("", "", ""),
        ("VERDICT BREAKDOWN", "", ""),
        ("PASS", passed, ""),
        ("WARN", warned, ""),
        ("FAIL", failed, ""),
        ("", "", ""),
        ("ISSUES BREAKDOWN", "", ""),
        ("High Severity Issues (Total)", total_high, "PASS" if total_high == 0 else "FAIL"),
        ("Medium Severity Issues (Total)", total_med, ""),
        ("Low Severity Issues (Total)", total_low, ""),
        ("", "", ""),
        ("AVERAGE SCORES", "PRODUCTION", "STAGING"),
        ("Compliance Score", avg_field(all_rows, "PROD_Compliance Score"), avg_field(all_rows, "STAGING_Compliance Score")),
        ("Sentiment Score", avg_field(all_rows, "PROD_Sentiment Score"), avg_field(all_rows, "STAGING_Sentiment Score")),
        ("Confidence Score", avg_field(all_rows, "PROD_Confidence Score"), avg_field(all_rows, "STAGING_Confidence Score")),
        ("BANT Overall", avg_field(all_rows, "PROD_BANT Overall"), avg_field(all_rows, "STAGING_BANT Overall")),
        ("Lead Score", avg_field(all_rows, "PROD_Lead Score"), avg_field(all_rows, "STAGING_Lead Score")),
        ("", "", ""),
        ("MATCH RATES", "COUNT", "RATE"),
        ("Booking Status Match", sum(1 for r in all_rows if r.get("match_booking") == "MATCH"),
         f"{sum(1 for r in all_rows if r.get('match_booking') == 'MATCH')}/{sum(1 for r in all_rows if r.get('match_booking') in ('MATCH','MISMATCH'))}"),
        ("Call Type Match", sum(1 for r in all_rows if r.get("match_call_type") == "MATCH"),
         f"{sum(1 for r in all_rows if r.get('match_call_type') == 'MATCH')}/{sum(1 for r in all_rows if r.get('match_call_type') in ('MATCH','MISMATCH'))}"),
        ("", "", ""),
        ("RECOMMENDATION", "", ""),
    ]

    # Determine recommendation
    if failed == 0 and total_high == 0:
        recommendation = "GO — Gemini outputs are at parity with OpenAI production"
    elif failed <= 2 and total_high <= 3:
        recommendation = "GO WITH CAVEATS — Minor differences noted, review flagged calls"
    else:
        recommendation = "NO-GO — Significant quality issues found, needs investigation"

    summary_data.append((recommendation, "", ""))

    # Write summary sheet
    for row_idx, (a, b, c) in enumerate(summary_data, 1):
        ws2.cell(row=row_idx, column=1, value=a).border = THIN_BORDER
        ws2.cell(row=row_idx, column=2, value=b).border = THIN_BORDER
        ws2.cell(row=row_idx, column=3, value=c).border = THIN_BORDER

    # Style header rows
    for row_idx, (a, b, c) in enumerate(summary_data, 1):
        if a in ("METRIC", "VERDICT BREAKDOWN", "ISSUES BREAKDOWN", "AVERAGE SCORES",
                 "MATCH RATES", "RECOMMENDATION"):
            for col in range(1, 4):
                cell = ws2.cell(row=row_idx, column=col)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL_META
        if a == "GEMINI MIGRATION — EXECUTIVE SUMMARY":
            cell = ws2.cell(row=row_idx, column=1)
            cell.font = Font(bold=True, size=14, color="2F5496")

    # Color-code status cells
    for row_idx, (a, b, c) in enumerate(summary_data, 1):
        status_cell = ws2.cell(row=row_idx, column=3)
        if c == "PASS":
            status_cell.fill = SEVERITY_FILLS["Low"]
            status_cell.font = Font(bold=True, color="006100")
        elif c == "FAIL":
            status_cell.fill = SEVERITY_FILLS["High"]
            status_cell.font = Font(bold=True, color="9C0006")
        elif c == "WARN":
            status_cell.fill = SEVERITY_FILLS["Medium"]
            status_cell.font = Font(bold=True, color="9C6500")

    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 25
    ws2.column_dimensions["C"].width = 25

    # ══════════════════════════════════════════════════════════════════════
    # Sheet 3: Issues Log
    # ══════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Issues Log")
    ws3.freeze_panes = "A2"

    issue_headers = ["Call #", "Audio URL", "Staging Call ID", "Severity", "Field", "Type", "Detail"]
    for col_idx, h in enumerate(issue_headers, 1):
        cell = ws3.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL_COMPARE
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    issue_row = 2
    for row_data in all_rows:
        issues_text = row_data.get("issues_detail", "")
        if not issues_text:
            continue
        for line in issues_text.split("\n"):
            if not line.strip():
                continue
            # Parse: [Severity] Field: Type — Detail
            parts = line.strip()
            severity = ""
            field = ""
            issue_type = ""
            detail = parts
            if parts.startswith("["):
                end = parts.find("]")
                if end > 0:
                    severity = parts[1:end]
                    rest = parts[end+2:]
                    if ": " in rest:
                        field, rest2 = rest.split(": ", 1)
                        if " — " in rest2:
                            issue_type, detail = rest2.split(" — ", 1)
                        else:
                            detail = rest2

            ws3.cell(row=issue_row, column=1, value=row_data.get("call_num")).border = THIN_BORDER
            ws3.cell(row=issue_row, column=2, value=row_data.get("audio_url")).border = THIN_BORDER
            ws3.cell(row=issue_row, column=3, value=row_data.get("staging_call_id")).border = THIN_BORDER
            sev_cell = ws3.cell(row=issue_row, column=4, value=severity)
            sev_cell.border = THIN_BORDER
            if severity in SEVERITY_FILLS:
                sev_cell.fill = SEVERITY_FILLS[severity]
            ws3.cell(row=issue_row, column=5, value=field).border = THIN_BORDER
            ws3.cell(row=issue_row, column=6, value=issue_type).border = THIN_BORDER
            ws3.cell(row=issue_row, column=7, value=detail).border = THIN_BORDER
            issue_row += 1

    ws3.column_dimensions["A"].width = 8
    ws3.column_dimensions["B"].width = 40
    ws3.column_dimensions["C"].width = 30
    ws3.column_dimensions["D"].width = 12
    ws3.column_dimensions["E"].width = 25
    ws3.column_dimensions["F"].width = 20
    ws3.column_dimensions["G"].width = 60

    wb.save(output_path)
    print(f"\nExcel report saved: {output_path}")


# ─────────────────────────────────────────────────────────────────────────────
# HTML Dashboard Generator
# ─────────────────────────────────────────────────────────────────────────────

import html as html_mod

def esc(text):
    if text is None:
        return ""
    return html_mod.escape(str(text))


def score_bar(score, label="", max_val=1.0):
    """Generate an inline score bar HTML snippet."""
    if score is None or score == "":
        return '<span class="na">N/A</span>'
    try:
        val = float(score)
    except (ValueError, TypeError):
        return f'<span class="na">{esc(str(score))}</span>'
    if max_val > 1:
        pct = min(int(val / max_val * 100), 100)
    else:
        pct = min(int(val * 100), 100) if val <= 1 else min(int(val), 100)
    if pct >= 80:
        color = "#22c55e"
    elif pct >= 60:
        color = "#eab308"
    elif pct >= 40:
        color = "#f97316"
    else:
        color = "#ef4444"
    display = f"{val:.2f}" if max_val <= 1 else f"{val:.0f}"
    return (
        f'<div class="score-bar">'
        f'<div class="score-fill" style="width:{pct}%;background:{color}"></div>'
        f'<span class="score-label">{display}</span>'
        f'</div>'
    )


def delta_badge(prod_row, staging_row, field, threshold=0.15, is_int=False):
    """Generate a delta badge comparing prod vs staging."""
    pv = prod_row.get(f"PROD_{field}")
    sv = staging_row.get(f"STAGING_{field}")
    if pv is None or sv is None or pv == "" or sv == "":
        return '<span class="badge badge-na">N/A</span>'
    try:
        pf = float(pv)
        sf = float(sv)
        d = abs(sf - pf)
        sign = "+" if sf >= pf else "-"
        display = f"{d:.0f}" if is_int else f"{d:.3f}"
        if d > threshold:
            return f'<span class="badge badge-fail">{sign}{display}</span>'
        elif d > threshold * 0.5:
            return f'<span class="badge badge-warn">{sign}{display}</span>'
        else:
            return f'<span class="badge badge-pass">{sign}{display}</span>'
    except (ValueError, TypeError):
        return '<span class="badge badge-na">N/A</span>'


def match_badge(prod_row, staging_row, field):
    """Generate a match/mismatch badge."""
    pv = str(prod_row.get(f"PROD_{field}", "")).strip()
    sv = str(staging_row.get(f"STAGING_{field}", "")).strip()
    if not pv and not sv:
        return '<span class="badge badge-na">N/A</span>'
    if pv.lower() == sv.lower():
        return '<span class="badge badge-pass">MATCH</span>'
    return f'<span class="badge badge-fail">MISMATCH</span>'


def write_html_dashboard(all_rows, all_comparison_data, output_path):
    """Generate a stakeholder-friendly HTML dashboard."""
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    total = len(all_rows)
    passed = sum(1 for r in all_rows if r.get("verdict") == "PASS")
    warned = sum(1 for r in all_rows if r.get("verdict") == "WARN")
    failed = sum(1 for r in all_rows if r.get("verdict") == "FAIL")
    staging_ok = sum(1 for r in all_rows if r.get("staging_process_status") == "completed")
    prod_found = sum(1 for r in all_rows if r.get("prod_call_id", "NOT FOUND") != "NOT FOUND")
    total_high = sum(r.get("high_issues", 0) for r in all_rows)
    total_med = sum(r.get("medium_issues", 0) for r in all_rows)
    total_low = sum(r.get("low_issues", 0) for r in all_rows)

    if failed == 0 and total_high == 0:
        recommendation = "GO"
        rec_class = "rec-go"
        rec_text = "Gemini outputs are at parity with OpenAI production"
    elif failed <= 2 and total_high <= 3:
        recommendation = "GO WITH CAVEATS"
        rec_class = "rec-caveat"
        rec_text = "Minor differences noted — review flagged calls before production rollout"
    else:
        recommendation = "NO-GO"
        rec_class = "rec-nogo"
        rec_text = "Significant quality issues found — needs investigation before production"

    def avg_field(key):
        vals = []
        for r in all_rows:
            v = r.get(key)
            if v is not None and v != "":
                try:
                    vals.append(float(v))
                except (ValueError, TypeError):
                    pass
        return round(sum(vals) / len(vals), 3) if vals else None

    # Build call cards HTML
    call_cards_html = []
    for idx, row in enumerate(all_rows):
        cd = all_comparison_data[idx] if idx < len(all_comparison_data) else {}
        prod_data = cd.get("prod_row", {})
        staging_data = cd.get("staging_row", {})
        issues = cd.get("issues", [])

        verdict = row.get("verdict", "")
        v_class = {"PASS": "verdict-pass", "WARN": "verdict-warn", "FAIL": "verdict-fail"}.get(verdict, "")

        # Failure reasons block
        issues_html = ""
        if issues:
            issues_html = '<div class="issues-block"><h4>Issues Found</h4><table class="issues-table"><tr><th>Severity</th><th>Field</th><th>Type</th><th>Detail</th></tr>'
            for iss in sorted(issues, key=lambda x: {"High": 0, "Medium": 1, "Low": 2}.get(x["severity"], 3)):
                sev_class = {"High": "sev-high", "Medium": "sev-med", "Low": "sev-low"}.get(iss["severity"], "")
                issues_html += (
                    f'<tr class="{sev_class}">'
                    f'<td><span class="sev-badge {sev_class}">{esc(iss["severity"])}</span></td>'
                    f'<td>{esc(iss["field"])}</td>'
                    f'<td>{esc(iss["type"])}</td>'
                    f'<td>{esc(iss["detail"])}</td>'
                    f'</tr>'
                )
            issues_html += '</table></div>'
        else:
            issues_html = '<div class="no-issues">No issues found</div>'

        # Determine batch label
        call_num = row.get("call_num", idx + 1)
        batch = row.get("source_batch", "")

        card = f'''
        <div class="call-card {v_class}">
            <div class="call-header" onclick="toggleCard(this)">
                <div class="call-title">
                    <span class="call-num">#{call_num}</span>
                    <span class="verdict-tag {v_class}">{verdict}</span>
                    <span class="batch-tag">{esc(batch)}</span>
                </div>
                <div class="call-meta">
                    <span>Staging: <strong>{esc(str(row.get("staging_process_status", "")))}</strong></span>
                    <span>Issues: <strong class="{"issue-count-bad" if row.get("high_issues", 0) > 0 else "issue-count-ok"}">{row.get("issues_count", 0)}</strong></span>
                    <span class="expand-icon">&#9660;</span>
                </div>
            </div>
            <div class="call-body" style="display:none">
                <div class="meta-row">
                    <div class="meta-item"><label>Audio URL</label><span class="audio-url">{esc(row.get("audio_url", ""))}</span></div>
                    <div class="meta-item"><label>PROD Call ID</label><span>{esc(str(row.get("prod_call_id", "NOT FOUND")))}</span></div>
                    <div class="meta-item"><label>STAGING Call ID</label><span>{esc(str(row.get("staging_call_id", "")))}</span></div>
                    <div class="meta-item"><label>Processing Time</label><span>{row.get("staging_process_time", "")}s</span></div>
                </div>

                <div class="comparison-grid">
                    <div class="comp-section">
                        <h4>Summary & Extraction</h4>
                        <table class="comp-table">
                            <tr><th>Field</th><th>Production (OpenAI)</th><th>Staging (Gemini)</th><th>Delta</th></tr>
                            <tr>
                                <td>Compliance Score</td>
                                <td>{score_bar(prod_data.get("PROD_Compliance Score"))}</td>
                                <td>{score_bar(staging_data.get("STAGING_Compliance Score"))}</td>
                                <td>{delta_badge(prod_data, staging_data, "Compliance Score", 0.15)}</td>
                            </tr>
                            <tr>
                                <td>Sentiment Score</td>
                                <td>{score_bar(prod_data.get("PROD_Sentiment Score"))}</td>
                                <td>{score_bar(staging_data.get("STAGING_Sentiment Score"))}</td>
                                <td>{delta_badge(prod_data, staging_data, "Sentiment Score", 0.3)}</td>
                            </tr>
                            <tr>
                                <td>Confidence Score</td>
                                <td>{score_bar(prod_data.get("PROD_Confidence Score"))}</td>
                                <td>{score_bar(staging_data.get("STAGING_Confidence Score"))}</td>
                                <td>{delta_badge(prod_data, staging_data, "Confidence Score", 0.2)}</td>
                            </tr>
                            <tr>
                                <td>Lead Score</td>
                                <td>{score_bar(prod_data.get("PROD_Lead Score"), max_val=100)}</td>
                                <td>{score_bar(staging_data.get("STAGING_Lead Score"), max_val=100)}</td>
                                <td>{delta_badge(prod_data, staging_data, "Lead Score", 15, is_int=True)}</td>
                            </tr>
                            <tr>
                                <td>Compliance Rate</td>
                                <td>{score_bar(prod_data.get("PROD_Compliance Rate"))}</td>
                                <td>{score_bar(staging_data.get("STAGING_Compliance Rate"))}</td>
                                <td>{delta_badge(prod_data, staging_data, "Compliance Rate", 0.15)}</td>
                            </tr>
                        </table>
                    </div>

                    <div class="comp-section">
                        <h4>BANT Scores</h4>
                        <table class="comp-table">
                            <tr><th>Dimension</th><th>Production</th><th>Staging</th><th>Delta</th></tr>
                            <tr>
                                <td>Budget</td>
                                <td>{score_bar(prod_data.get("PROD_BANT Budget"))}</td>
                                <td>{score_bar(staging_data.get("STAGING_BANT Budget"))}</td>
                                <td>{delta_badge(prod_data, staging_data, "BANT Budget", 0.2)}</td>
                            </tr>
                            <tr>
                                <td>Authority</td>
                                <td>{score_bar(prod_data.get("PROD_BANT Authority"))}</td>
                                <td>{score_bar(staging_data.get("STAGING_BANT Authority"))}</td>
                                <td>{delta_badge(prod_data, staging_data, "BANT Authority", 0.2)}</td>
                            </tr>
                            <tr>
                                <td>Need</td>
                                <td>{score_bar(prod_data.get("PROD_BANT Need"))}</td>
                                <td>{score_bar(staging_data.get("STAGING_BANT Need"))}</td>
                                <td>{delta_badge(prod_data, staging_data, "BANT Need", 0.2)}</td>
                            </tr>
                            <tr>
                                <td>Timeline</td>
                                <td>{score_bar(prod_data.get("PROD_BANT Timeline"))}</td>
                                <td>{score_bar(staging_data.get("STAGING_BANT Timeline"))}</td>
                                <td>{delta_badge(prod_data, staging_data, "BANT Timeline", 0.2)}</td>
                            </tr>
                            <tr>
                                <td><strong>Overall</strong></td>
                                <td>{score_bar(prod_data.get("PROD_BANT Overall"))}</td>
                                <td>{score_bar(staging_data.get("STAGING_BANT Overall"))}</td>
                                <td>{delta_badge(prod_data, staging_data, "BANT Overall", 0.2)}</td>
                            </tr>
                        </table>
                    </div>

                    <div class="comp-section">
                        <h4>Classification Matching</h4>
                        <table class="comp-table">
                            <tr><th>Field</th><th>Production</th><th>Staging</th><th>Match</th></tr>
                            <tr>
                                <td>Booking Status</td>
                                <td>{esc(str(prod_data.get("PROD_Booking Status", "")))}</td>
                                <td>{esc(str(staging_data.get("STAGING_Booking Status", "")))}</td>
                                <td>{match_badge(prod_data, staging_data, "Booking Status")}</td>
                            </tr>
                            <tr>
                                <td>Call Type</td>
                                <td>{esc(str(prod_data.get("PROD_Call Type", "")))}</td>
                                <td>{esc(str(staging_data.get("STAGING_Call Type", "")))}</td>
                                <td>{match_badge(prod_data, staging_data, "Call Type")}</td>
                            </tr>
                            <tr>
                                <td>Qualification Status</td>
                                <td>{esc(str(prod_data.get("PROD_Qualification Status", "")))}</td>
                                <td>{esc(str(staging_data.get("STAGING_Qualification Status", "")))}</td>
                                <td>{match_badge(prod_data, staging_data, "Qualification Status")}</td>
                            </tr>
                            <tr>
                                <td>Call Outcome</td>
                                <td>{esc(str(prod_data.get("PROD_Call Outcome", "")))}</td>
                                <td>{esc(str(staging_data.get("STAGING_Call Outcome", "")))}</td>
                                <td>{match_badge(prod_data, staging_data, "Call Outcome")}</td>
                            </tr>
                            <tr>
                                <td>Scope Classification</td>
                                <td>{esc(str(prod_data.get("PROD_Scope Classification", "")))}</td>
                                <td>{esc(str(staging_data.get("STAGING_Scope Classification", "")))}</td>
                                <td>{match_badge(prod_data, staging_data, "Scope Classification")}</td>
                            </tr>
                            <tr>
                                <td>Customer Name</td>
                                <td>{esc(str(prod_data.get("PROD_Customer Name", "")))}</td>
                                <td>{esc(str(staging_data.get("STAGING_Customer Name", "")))}</td>
                                <td>{match_badge(prod_data, staging_data, "Customer Name")}</td>
                            </tr>
                        </table>
                    </div>

                    <div class="comp-section">
                        <h4>Objections & Compliance</h4>
                        <table class="comp-table">
                            <tr><th>Field</th><th>Production</th><th>Staging</th><th>Delta</th></tr>
                            <tr>
                                <td>Objections Count</td>
                                <td>{esc(str(prod_data.get("PROD_Objections Count", "")))}</td>
                                <td>{esc(str(staging_data.get("STAGING_Objections Count", "")))}</td>
                                <td>{delta_badge(prod_data, staging_data, "Objections Count", 2, is_int=True)}</td>
                            </tr>
                            <tr>
                                <td>Stages Followed</td>
                                <td class="small-text">{esc(str(prod_data.get("PROD_Stages Followed", "")))}</td>
                                <td class="small-text">{esc(str(staging_data.get("STAGING_Stages Followed", "")))}</td>
                                <td></td>
                            </tr>
                            <tr>
                                <td>Stages Missed</td>
                                <td class="small-text">{esc(str(prod_data.get("PROD_Stages Missed", "")))}</td>
                                <td class="small-text">{esc(str(staging_data.get("STAGING_Stages Missed", "")))}</td>
                                <td></td>
                            </tr>
                            <tr>
                                <td>Transcript Length</td>
                                <td>{esc(str(prod_data.get("PROD_Transcript Length", "")))}</td>
                                <td>{esc(str(staging_data.get("STAGING_Transcript Length", "")))}</td>
                                <td></td>
                            </tr>
                            <tr>
                                <td>Segment Count</td>
                                <td>{esc(str(prod_data.get("PROD_Segment Count", "")))}</td>
                                <td>{esc(str(staging_data.get("STAGING_Segment Count", "")))}</td>
                                <td></td>
                            </tr>
                        </table>
                    </div>

                    <div class="comp-section full-width">
                        <h4>Summary Comparison</h4>
                        <div class="summary-compare">
                            <div class="summary-col">
                                <h5>Production (OpenAI)</h5>
                                <p>{esc(str(prod_data.get("PROD_Summary", "N/A")))[:500]}{"..." if len(str(prod_data.get("PROD_Summary", ""))) > 500 else ""}</p>
                            </div>
                            <div class="summary-col">
                                <h5>Staging (Gemini)</h5>
                                <p>{esc(str(staging_data.get("STAGING_Summary", "N/A")))[:500]}{"..." if len(str(staging_data.get("STAGING_Summary", ""))) > 500 else ""}</p>
                            </div>
                        </div>
                    </div>
                </div>

                {issues_html}
            </div>
        </div>
        '''
        call_cards_html.append(card)

    # Aggregate match rates
    booking_match = sum(1 for r in all_rows if r.get("match_booking") == "MATCH")
    booking_total = sum(1 for r in all_rows if r.get("match_booking") in ("MATCH", "MISMATCH"))
    calltype_match = sum(1 for r in all_rows if r.get("match_call_type") == "MATCH")
    calltype_total = sum(1 for r in all_rows if r.get("match_call_type") in ("MATCH", "MISMATCH"))

    html = f'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Otto Intelligence — Gemini vs Production Comparison Dashboard</title>
<style>
    :root {{
        --bg: #0f172a; --surface: #1e293b; --surface2: #334155;
        --border: #475569; --text: #e2e8f0; --text-dim: #94a3b8;
        --green: #22c55e; --yellow: #eab308; --red: #ef4444; --orange: #f97316;
        --blue: #3b82f6; --purple: #8b5cf6;
    }}
    * {{ margin: 0; padding: 0; box-sizing: border-box; }}
    body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; background: var(--bg); color: var(--text); line-height: 1.5; }}

    .container {{ max-width: 1400px; margin: 0 auto; padding: 24px; }}

    /* Header */
    .header {{ text-align: center; padding: 40px 20px 30px; border-bottom: 1px solid var(--border); margin-bottom: 30px; }}
    .header h1 {{ font-size: 28px; font-weight: 700; margin-bottom: 4px; }}
    .header h1 span {{ color: var(--blue); }}
    .header .subtitle {{ color: var(--text-dim); font-size: 14px; }}

    /* Recommendation Banner */
    .rec-banner {{ padding: 20px 30px; border-radius: 12px; margin-bottom: 30px; text-align: center; }}
    .rec-go {{ background: linear-gradient(135deg, #065f46, #064e3b); border: 1px solid var(--green); }}
    .rec-caveat {{ background: linear-gradient(135deg, #713f12, #78350f); border: 1px solid var(--yellow); }}
    .rec-nogo {{ background: linear-gradient(135deg, #7f1d1d, #991b1b); border: 1px solid var(--red); }}
    .rec-banner h2 {{ font-size: 24px; margin-bottom: 6px; }}
    .rec-banner p {{ color: var(--text-dim); font-size: 14px; }}

    /* KPI Grid */
    .kpi-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr)); gap: 16px; margin-bottom: 30px; }}
    .kpi-card {{ background: var(--surface); border: 1px solid var(--border); border-radius: 10px; padding: 20px; text-align: center; }}
    .kpi-card .kpi-value {{ font-size: 32px; font-weight: 700; }}
    .kpi-card .kpi-label {{ font-size: 12px; color: var(--text-dim); text-transform: uppercase; letter-spacing: 0.5px; margin-top: 4px; }}
    .kpi-card .kpi-sub {{ font-size: 11px; color: var(--text-dim); margin-top: 2px; }}
    .kpi-green {{ color: var(--green); }}
    .kpi-yellow {{ color: var(--yellow); }}
    .kpi-red {{ color: var(--red); }}
    .kpi-blue {{ color: var(--blue); }}

    /* Score Averages */
    .avg-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(280px, 1fr)); gap: 16px; margin-bottom: 30px; }}
    .avg-card {{ background: var(--surface); border: 1px solid var(--border); border-radius: 10px; padding: 20px; }}
    .avg-card h3 {{ font-size: 14px; color: var(--text-dim); margin-bottom: 12px; text-transform: uppercase; letter-spacing: 0.5px; }}
    .avg-row {{ display: flex; justify-content: space-between; align-items: center; padding: 8px 0; border-bottom: 1px solid var(--surface2); }}
    .avg-row:last-child {{ border-bottom: none; }}
    .avg-row .label {{ font-size: 13px; }}
    .avg-row .values {{ display: flex; gap: 16px; align-items: center; font-size: 13px; }}
    .avg-row .prod-val {{ color: var(--green); font-weight: 600; }}
    .avg-row .stage-val {{ color: var(--yellow); font-weight: 600; }}

    /* Filters */
    .filters {{ display: flex; gap: 10px; margin-bottom: 20px; flex-wrap: wrap; }}
    .filter-btn {{ padding: 8px 16px; border-radius: 8px; border: 1px solid var(--border); background: var(--surface); color: var(--text); cursor: pointer; font-size: 13px; transition: all 0.2s; }}
    .filter-btn:hover, .filter-btn.active {{ background: var(--blue); border-color: var(--blue); color: white; }}

    /* Call Cards */
    .call-card {{ background: var(--surface); border: 1px solid var(--border); border-radius: 10px; margin-bottom: 12px; overflow: hidden; transition: all 0.2s; }}
    .call-card.verdict-fail {{ border-left: 4px solid var(--red); }}
    .call-card.verdict-warn {{ border-left: 4px solid var(--yellow); }}
    .call-card.verdict-pass {{ border-left: 4px solid var(--green); }}

    .call-header {{ display: flex; justify-content: space-between; align-items: center; padding: 14px 20px; cursor: pointer; }}
    .call-header:hover {{ background: var(--surface2); }}
    .call-title {{ display: flex; align-items: center; gap: 10px; }}
    .call-num {{ font-weight: 700; font-size: 15px; color: var(--text-dim); }}
    .call-meta {{ display: flex; align-items: center; gap: 16px; font-size: 13px; color: var(--text-dim); }}
    .expand-icon {{ font-size: 12px; transition: transform 0.2s; }}
    .call-card.open .expand-icon {{ transform: rotate(180deg); }}

    .verdict-tag {{ padding: 3px 10px; border-radius: 6px; font-size: 11px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.5px; }}
    .verdict-tag.verdict-pass {{ background: #065f4620; color: var(--green); border: 1px solid var(--green); }}
    .verdict-tag.verdict-warn {{ background: #78350f20; color: var(--yellow); border: 1px solid var(--yellow); }}
    .verdict-tag.verdict-fail {{ background: #7f1d1d20; color: var(--red); border: 1px solid var(--red); }}

    .batch-tag {{ padding: 2px 8px; border-radius: 4px; font-size: 11px; background: var(--surface2); color: var(--text-dim); }}

    .issue-count-bad {{ color: var(--red); }}
    .issue-count-ok {{ color: var(--green); }}

    .call-body {{ padding: 0 20px 20px; }}
    .meta-row {{ display: flex; flex-wrap: wrap; gap: 16px; margin-bottom: 16px; padding: 12px; background: var(--surface2); border-radius: 8px; }}
    .meta-item {{ flex: 1; min-width: 200px; }}
    .meta-item label {{ display: block; font-size: 11px; color: var(--text-dim); text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 2px; }}
    .meta-item span {{ font-size: 13px; word-break: break-all; }}
    .audio-url {{ font-family: monospace; font-size: 11px !important; }}

    /* Comparison Grid */
    .comparison-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 16px; margin-bottom: 16px; }}
    .comp-section {{ background: var(--bg); border-radius: 8px; padding: 16px; }}
    .comp-section.full-width {{ grid-column: 1 / -1; }}
    .comp-section h4 {{ font-size: 13px; color: var(--text-dim); text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 10px; }}

    .comp-table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
    .comp-table th {{ text-align: left; padding: 6px 8px; color: var(--text-dim); font-size: 11px; text-transform: uppercase; border-bottom: 1px solid var(--surface2); }}
    .comp-table td {{ padding: 6px 8px; border-bottom: 1px solid var(--surface2); vertical-align: middle; }}
    .small-text {{ font-size: 11px; max-width: 200px; word-wrap: break-word; }}

    /* Score bars */
    .score-bar {{ position: relative; width: 100px; height: 22px; background: var(--surface2); border-radius: 4px; overflow: hidden; }}
    .score-fill {{ position: absolute; top: 0; left: 0; height: 100%; border-radius: 4px; transition: width 0.3s; }}
    .score-label {{ position: absolute; top: 50%; left: 50%; transform: translate(-50%, -50%); font-size: 11px; font-weight: 700; color: white; text-shadow: 0 1px 2px rgba(0,0,0,0.5); }}
    .na {{ color: var(--text-dim); font-size: 12px; }}

    /* Badges */
    .badge {{ padding: 2px 8px; border-radius: 4px; font-size: 11px; font-weight: 600; }}
    .badge-pass {{ background: #065f4640; color: var(--green); }}
    .badge-warn {{ background: #78350f40; color: var(--yellow); }}
    .badge-fail {{ background: #7f1d1d40; color: var(--red); }}
    .badge-na {{ background: var(--surface2); color: var(--text-dim); }}

    /* Issues */
    .issues-block {{ margin-top: 12px; }}
    .issues-block h4 {{ font-size: 13px; color: var(--red); margin-bottom: 8px; }}
    .issues-table {{ width: 100%; border-collapse: collapse; font-size: 12px; }}
    .issues-table th {{ text-align: left; padding: 6px 10px; background: var(--bg); color: var(--text-dim); font-size: 11px; text-transform: uppercase; }}
    .issues-table td {{ padding: 6px 10px; border-bottom: 1px solid var(--surface2); }}
    .sev-badge {{ padding: 2px 8px; border-radius: 4px; font-size: 11px; font-weight: 700; }}
    .sev-high {{ }}
    .sev-high .sev-badge {{ background: #7f1d1d; color: #fca5a5; }}
    .sev-med .sev-badge {{ background: #78350f; color: #fde68a; }}
    .sev-low .sev-badge {{ background: #065f46; color: #bbf7d0; }}
    .no-issues {{ padding: 12px; text-align: center; color: var(--green); font-size: 13px; background: #065f4615; border-radius: 8px; }}

    /* Summary comparison */
    .summary-compare {{ display: grid; grid-template-columns: 1fr 1fr; gap: 16px; }}
    .summary-col {{ padding: 12px; background: var(--surface); border-radius: 8px; }}
    .summary-col h5 {{ font-size: 12px; color: var(--text-dim); margin-bottom: 8px; text-transform: uppercase; }}
    .summary-col p {{ font-size: 12px; line-height: 1.6; color: var(--text); }}

    /* Footer */
    .footer {{ text-align: center; padding: 30px; color: var(--text-dim); font-size: 12px; border-top: 1px solid var(--border); margin-top: 40px; }}

    @media (max-width: 768px) {{
        .comparison-grid {{ grid-template-columns: 1fr; }}
        .summary-compare {{ grid-template-columns: 1fr; }}
        .kpi-grid {{ grid-template-columns: repeat(2, 1fr); }}
    }}
</style>
</head>
<body>
<div class="container">

    <div class="header">
        <h1>Otto Intelligence — <span>Gemini vs Production</span></h1>
        <div class="subtitle">Comparison Dashboard | Generated {now} | {total} Calls Tested</div>
    </div>

    <div class="rec-banner {rec_class}">
        <h2>{recommendation}</h2>
        <p>{rec_text}</p>
    </div>

    <div class="kpi-grid">
        <div class="kpi-card">
            <div class="kpi-value kpi-blue">{total}</div>
            <div class="kpi-label">Total Calls</div>
        </div>
        <div class="kpi-card">
            <div class="kpi-value kpi-green">{passed}</div>
            <div class="kpi-label">Passed</div>
        </div>
        <div class="kpi-card">
            <div class="kpi-value kpi-yellow">{warned}</div>
            <div class="kpi-label">Warnings</div>
        </div>
        <div class="kpi-card">
            <div class="kpi-value kpi-red">{failed}</div>
            <div class="kpi-label">Failed</div>
        </div>
        <div class="kpi-card">
            <div class="kpi-value kpi-green">{staging_ok}/{total}</div>
            <div class="kpi-label">Staging Processed</div>
        </div>
        <div class="kpi-card">
            <div class="kpi-value kpi-blue">{prod_found}/{total}</div>
            <div class="kpi-label">Prod Baseline Found</div>
        </div>
        <div class="kpi-card">
            <div class="kpi-value kpi-red">{total_high}</div>
            <div class="kpi-label">High Issues</div>
        </div>
        <div class="kpi-card">
            <div class="kpi-value kpi-yellow">{total_med}</div>
            <div class="kpi-label">Medium Issues</div>
        </div>
    </div>

    <div class="avg-grid">
        <div class="avg-card">
            <h3>Average Scores Comparison</h3>
            <div class="avg-row"><span class="label">Compliance Score</span><div class="values"><span class="prod-val">{avg_field("PROD_Compliance Score") or "N/A"}</span><span class="stage-val">{avg_field("STAGING_Compliance Score") or "N/A"}</span></div></div>
            <div class="avg-row"><span class="label">Sentiment Score</span><div class="values"><span class="prod-val">{avg_field("PROD_Sentiment Score") or "N/A"}</span><span class="stage-val">{avg_field("STAGING_Sentiment Score") or "N/A"}</span></div></div>
            <div class="avg-row"><span class="label">Confidence Score</span><div class="values"><span class="prod-val">{avg_field("PROD_Confidence Score") or "N/A"}</span><span class="stage-val">{avg_field("STAGING_Confidence Score") or "N/A"}</span></div></div>
            <div class="avg-row"><span class="label">BANT Overall</span><div class="values"><span class="prod-val">{avg_field("PROD_BANT Overall") or "N/A"}</span><span class="stage-val">{avg_field("STAGING_BANT Overall") or "N/A"}</span></div></div>
            <div class="avg-row"><span class="label">Lead Score</span><div class="values"><span class="prod-val">{avg_field("PROD_Lead Score") or "N/A"}</span><span class="stage-val">{avg_field("STAGING_Lead Score") or "N/A"}</span></div></div>
            <div class="avg-row" style="margin-top:8px;border-top:1px solid var(--border);padding-top:8px;font-size:11px;color:var(--text-dim)"><span>Legend:</span><div class="values"><span class="prod-val">Production</span><span class="stage-val">Staging</span></div></div>
        </div>
        <div class="avg-card">
            <h3>Match Rates</h3>
            <div class="avg-row"><span class="label">Booking Status</span><div class="values"><strong>{"%.0f" % (booking_match/booking_total*100) if booking_total > 0 else "N/A"}%</strong><span style="color:var(--text-dim)">{booking_match}/{booking_total}</span></div></div>
            <div class="avg-row"><span class="label">Call Type</span><div class="values"><strong>{"%.0f" % (calltype_match/calltype_total*100) if calltype_total > 0 else "N/A"}%</strong><span style="color:var(--text-dim)">{calltype_match}/{calltype_total}</span></div></div>
            <div class="avg-row"><span class="label" style="margin-top:16px;font-weight:600">Issue Breakdown</span><div class="values"></div></div>
            <div class="avg-row"><span class="label">High Severity</span><div class="values"><span style="color:var(--red);font-weight:700">{total_high}</span></div></div>
            <div class="avg-row"><span class="label">Medium Severity</span><div class="values"><span style="color:var(--yellow);font-weight:700">{total_med}</span></div></div>
            <div class="avg-row"><span class="label">Low Severity</span><div class="values"><span style="color:var(--green);font-weight:700">{total_low}</span></div></div>
        </div>
    </div>

    <h2 style="font-size:18px;margin-bottom:16px">Per-Call Results</h2>

    <div class="filters">
        <button class="filter-btn active" onclick="filterCards('all')">All ({total})</button>
        <button class="filter-btn" onclick="filterCards('PASS')">Pass ({passed})</button>
        <button class="filter-btn" onclick="filterCards('WARN')">Warn ({warned})</button>
        <button class="filter-btn" onclick="filterCards('FAIL')">Fail ({failed})</button>
    </div>

    <div id="call-cards">
        {"".join(call_cards_html)}
    </div>

    <div class="footer">
        Otto Intelligence — Gemini Migration Comparison Report<br>
        Generated {now} | Staging: {STAGING_BASE_URL} | Production: {PROD_BASE_URL}
    </div>
</div>

<script>
function toggleCard(header) {{
    const card = header.closest('.call-card');
    const body = card.querySelector('.call-body');
    card.classList.toggle('open');
    body.style.display = body.style.display === 'none' ? 'block' : 'none';
}}

function filterCards(type) {{
    document.querySelectorAll('.filter-btn').forEach(b => b.classList.remove('active'));
    event.target.classList.add('active');
    document.querySelectorAll('.call-card').forEach(card => {{
        if (type === 'all') {{
            card.style.display = 'block';
        }} else {{
            const tag = card.querySelector('.verdict-tag');
            card.style.display = tag && tag.textContent.trim() === type ? 'block' : 'none';
        }}
    }});
}}

// Expand all failed cards by default
document.addEventListener('DOMContentLoaded', () => {{
    document.querySelectorAll('.call-card.verdict-fail').forEach(card => {{
        card.classList.add('open');
        card.querySelector('.call-body').style.display = 'block';
    }});
}});
</script>
</body>
</html>'''

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"HTML dashboard saved: {output_path}")


# ─────────────────────────────────────────────────────────────────────────────
# Main Orchestrator
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Otto Gemini vs Production Comparison Test")
    parser.add_argument("--skip-submit", action="store_true",
                        help="Skip submitting to staging; only fetch existing results and compare")
    parser.add_argument("--batch-size", type=int, default=5,
                        help="Number of calls to submit before waiting for completion (default: 5)")
    parser.add_argument("--output", type=str, default=EXCEL_OUTPUT,
                        help="Output Excel file path")
    args = parser.parse_args()

    print("=" * 70)
    print("  OTTO INTELLIGENCE — GEMINI vs PRODUCTION COMPARISON TEST")
    print(f"  Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  Audio URLs: {len(AUDIO_URLS)}")
    print(f"  Staging: {STAGING_BASE_URL}")
    print(f"  Production: {PROD_BASE_URL}")
    print("=" * 70)

    # ── Step 1: Health checks ──
    print("\n[1/5] Health Checks")
    staging_ok = health_check(STAGING_BASE_URL, STAGING_HEADERS, "Staging")
    prod_ok = health_check(PROD_BASE_URL, PROD_HEADERS, "Production")
    if not staging_ok:
        print("  WARNING: Staging is unreachable. Will skip staging submission.")
    if not prod_ok:
        print("  WARNING: Production is unreachable. Will skip production fetch.")

    all_results = []

    # ── Step 2: Find production baseline for each audio URL ──
    print("\n[2/5] Finding Production Baseline Calls")
    prod_call_map = {}  # audio_url -> prod call_id
    for i, audio_url in enumerate(AUDIO_URLS):
        print(f"  [{i+1:2d}/{len(AUDIO_URLS)}] Searching prod for: ...{audio_url[-40:]}")
        if prod_ok:
            prod_cid = find_prod_call_by_audio(audio_url)
            if prod_cid:
                prod_call_map[audio_url] = prod_cid
                print(f"         Found: {prod_cid}")
            else:
                print(f"         NOT FOUND in production")
        else:
            print(f"         SKIPPED (prod unreachable)")

    print(f"\n  Production baseline: {len(prod_call_map)}/{len(AUDIO_URLS)} calls found")

    # ── Step 3: Submit to staging (in batches) ──
    staging_call_map = {}  # audio_url -> (call_id, job_id)
    staging_results = {}  # audio_url -> (status, duration)

    if not args.skip_submit and staging_ok:
        print(f"\n[3/5] Submitting to Staging (batch_size={args.batch_size})")
        for batch_start in range(0, len(AUDIO_URLS), args.batch_size):
            batch = AUDIO_URLS[batch_start:batch_start + args.batch_size]
            batch_num = (batch_start // args.batch_size) + 1
            total_batches = (len(AUDIO_URLS) + args.batch_size - 1) // args.batch_size
            print(f"\n  --- Batch {batch_num}/{total_batches} ---")

            # Submit batch
            batch_jobs = []
            for i, audio_url in enumerate(batch):
                global_idx = batch_start + i
                call_id, job_id = submit_call(STAGING_BASE_URL, STAGING_HEADERS, audio_url, global_idx)
                staging_call_map[audio_url] = (call_id, job_id)
                batch_jobs.append((audio_url, call_id, job_id))

            # Poll batch
            print(f"\n  Polling batch {batch_num}...")
            for audio_url, call_id, job_id in batch_jobs:
                status, duration = poll_job(STAGING_BASE_URL, STAGING_HEADERS, job_id, f"[{call_id}]")
                staging_results[audio_url] = (status, duration)
                print(f"    {call_id}: {status} ({duration}s)")
    elif args.skip_submit:
        print("\n[3/5] SKIPPED (--skip-submit flag)")
    else:
        print("\n[3/5] SKIPPED (staging unreachable)")

    # ── Step 4: Fetch all data and compare ──
    print("\n[4/5] Fetching Data & Comparing")
    all_rows = []

    for i, audio_url in enumerate(AUDIO_URLS):
        print(f"\n  Call {i+1}/{len(AUDIO_URLS)}: ...{audio_url[-40:]}")

        # Production data
        prod_call_id = prod_call_map.get(audio_url)
        prod_row = {}
        if prod_call_id and prod_ok:
            print(f"    Fetching PROD data for {prod_call_id}...")
            prod_summary = fetch_summary(PROD_BASE_URL, PROD_HEADERS, prod_call_id)
            prod_detail = fetch_detail(PROD_BASE_URL, PROD_HEADERS, prod_call_id)
            prod_row = extract_row(prod_summary, prod_detail, prefix="PROD_")
        else:
            prod_row = {"PROD_Status": "NOT FOUND"}

        # Staging data
        staging_info = staging_call_map.get(audio_url, (None, None))
        staging_call_id = staging_info[0]
        staging_status_info = staging_results.get(audio_url, ("not_submitted", 0))
        staging_row = {}
        if staging_call_id and staging_ok:
            print(f"    Fetching STAGING data for {staging_call_id}...")
            staging_summary = fetch_summary(STAGING_BASE_URL, STAGING_HEADERS, staging_call_id)
            staging_detail = fetch_detail(STAGING_BASE_URL, STAGING_HEADERS, staging_call_id)
            staging_row = extract_row(staging_summary, staging_detail, prefix="STAGING_")
        else:
            staging_row = {"STAGING_Status": "NO DATA"}

        # Compare
        issues = compare_fields(prod_row, staging_row)
        high_count = sum(1 for iss in issues if iss["severity"] == "High")
        med_count = sum(1 for iss in issues if iss["severity"] == "Medium")
        print(f"    Issues: {len(issues)} (High={high_count}, Med={med_count})")

        # Build row
        row = build_comparison_row(
            i, audio_url, prod_call_id, staging_call_id,
            staging_status_info[0], staging_status_info[1],
            prod_row, staging_row, issues,
        )
        all_rows.append(row)

        # Store for JSON dump + HTML dashboard
        all_results.append({
            "audio_url": audio_url,
            "prod_call_id": prod_call_id,
            "staging_call_id": staging_call_id,
            "staging_status": staging_status_info[0],
            "staging_time": staging_status_info[1],
            "issues": issues,
            "verdict": row.get("verdict"),
            "prod_row": prod_row,
            "staging_row": staging_row,
        })

    # ── Step 5: Generate reports ──
    print("\n[5/5] Generating Reports")
    write_html_dashboard(all_rows, all_results, HTML_OUTPUT)
    write_excel(all_rows, args.output)

    # JSON dump (strip large nested dicts for cleaner JSON)
    json_results = []
    for r in all_results:
        json_results.append({
            "audio_url": r["audio_url"],
            "prod_call_id": r["prod_call_id"],
            "staging_call_id": r["staging_call_id"],
            "staging_status": r["staging_status"],
            "staging_time": r["staging_time"],
            "issues": r["issues"],
            "verdict": r["verdict"],
        })
    with open(JSON_OUTPUT, "w") as f:
        json.dump(json_results, f, indent=2, default=str)
    print(f"JSON data saved: {JSON_OUTPUT}")

    # Final summary
    total = len(all_rows)
    passed = sum(1 for r in all_rows if r.get("verdict") == "PASS")
    warned = sum(1 for r in all_rows if r.get("verdict") == "WARN")
    failed = sum(1 for r in all_rows if r.get("verdict") == "FAIL")

    print("\n" + "=" * 70)
    print("  FINAL SUMMARY")
    print("=" * 70)
    print(f"  Total Calls:   {total}")
    print(f"  PASS:          {passed}")
    print(f"  WARN:          {warned}")
    print(f"  FAIL:          {failed}")
    print(f"  Prod Found:    {len(prod_call_map)}/{total}")
    print(f"  Staging OK:    {sum(1 for v in staging_results.values() if v[0] == 'completed')}/{total}")
    print("=" * 70)

    if failed == 0:
        print("  RECOMMENDATION: GO")
    elif failed <= 2:
        print("  RECOMMENDATION: GO WITH CAVEATS")
    else:
        print("  RECOMMENDATION: NO-GO")
    print("=" * 70)


if __name__ == "__main__":
    main()
