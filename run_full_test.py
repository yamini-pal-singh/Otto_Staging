#!/usr/bin/env python3
"""
Otto Intelligence — Full Test: Submit to BOTH environments, fetch, compare.
Generates a 3-tab Excel matching the reference sheet format:
  Tab 1: Staging (Gemini) data
  Tab 2: Production (OpenAI) data
  Tab 3: Comparison (side-by-side with deltas + failure reasons)
"""
import os, sys, uuid, time, json, re
from datetime import datetime
from collections import defaultdict

import requests

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("pip install openpyxl"); sys.exit(1)

# ─── Config ───
STAGING_URL = "https://ottoai-stage.shunyalabs.ai"
PROD_URL    = "https://ottoai.shunyalabs.ai"
API_KEY     = "5q3fwliU9ZFo3epTCsUfUiDw1Dy4DnBP"
COMPANY_ID  = "1be5ea90-d3ae-4b03-8b05-f5679cd73bc4"
HEADERS     = {"X-API-Key": API_KEY, "Content-Type": "application/json"}
TIMEOUT     = 60
POLL_INTERVAL = 15
MAX_POLL     = 1800  # 30 min max per call

OUTPUT_DIR  = os.path.dirname(os.path.abspath(__file__))
EXCEL_OUT   = os.path.join(OUTPUT_DIR, "Gemini_vs_OpenAI_Full_Report.xlsx")
HTML_OUT    = os.path.join(OUTPUT_DIR, "gemini_comparison_dashboard.html")

AUDIO_URLS = [
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/bd24a254-fe4e-48ef-9c3e-3c3f3adfd68c/4082148782.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/00273c23-da2c-4e1b-b49c-4ff85d4a766a/4082112377.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/c0183543-60e4-434f-a247-89ec9ef8e1e3/4081871096.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/eb359c0a-02e2-4e94-b018-b3585c8a4024/4081765241.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/4da88a74-f129-481d-92d4-626c5a728835/4081742162.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/c6c4fff3-72d5-42d5-9cb4-16b8272e3d3b/4079556188.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/ddccb9c0-fce8-4eb5-81bd-5e738cb95979/4078581674.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/add722b1-c29f-4615-b389-4d29773d5958/st_309419607.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/df6c19b0-5d39-457b-8053-4dd83f4407cf/4076546354.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/6dd70da4-9d39-4496-871e-827af1429e29/4077402764.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/0caf5493-4ea1-4298-bee3-5ea9a83a2431/4094726633.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/31047617-96e2-4b71-8167-9b6aa3acaa3e/4094713955.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/cd5faebf-2c6b-4733-b560-e2a97eec644f/4094604863.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/90684b34-3972-4abe-a56b-07633b75da39/4094588411.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/349c746c-0a9b-44ff-a6c3-8c052a1efd29/4094548208.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/bdd04264-1591-4b4a-8aca-6f63beb79aba/4094484647.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/c979ee91-9189-4473-bcae-898ddd5d0028/4094457467.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/203c3c8b-4cce-4dfe-8ef7-74c4dd9ae3db/4094238074.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/1a6fa204-73b1-42e6-9adc-930a8be7196c/4093544600.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/cd8c80e2-0388-4ac7-818c-6f543bb500e8/4094042249.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/43ccc787-7f07-4a13-ac9d-d672d25a809c/4043504735.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/1fd7bea5-9ace-4e8f-a31f-152ea8269927/4015296617.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/bc946fa9-1e8a-4f1e-920c-64c4369fe778/4049722733.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/e2b0dab7-a05b-4448-8c13-fa3753f405ae/4036836500.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/8c6b15ee-5675-4e01-8b31-ff3658126353/4049722733.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/6e37c8bb-16bc-4e17-867e-ae5e9f57c3b9/4037028977.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/56dc7e30-ffed-4f8d-80eb-b514ffb30a50/4050591020.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/6566e3b9-acac-4b55-aad1-5742464107fa/4058579492.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/33cc8375-04ee-43a4-8644-bdffdb8d1b1b/4060192187.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/b9a3deca-d3e1-47ad-a2ad-875e58c0b7dc/4058565425.mp3",
]

# ─── API helpers ───

def submit(base, audio_url, idx, tag):
    call_id = f"{tag}_{idx+1}_{uuid.uuid4().hex[:8]}"
    payload = {
        "call_id": call_id, "company_id": COMPANY_ID, "audio_url": audio_url,
        "phone_number": "+14805551234", "rep_role": "customer_rep", "allow_reprocess": True,
        "metadata": {"agent": {"id": "USR_ANTHONY_ARIZONA", "name": "Anthony", "email": "anthony@arizonaroofers.com"}},
    }
    try:
        r = requests.post(f"{base}/api/v1/call-processing/process", headers=HEADERS, json=payload, timeout=TIMEOUT)
        d = r.json()
        print(f"  [{idx+1:2d}] HTTP {r.status_code} | {call_id} | job={d.get('job_id','N/A')}")
        return call_id, d.get("job_id")
    except Exception as e:
        print(f"  [{idx+1:2d}] ERROR: {e}")
        return call_id, None


def poll(base, job_id, label=""):
    if not job_id: return "no_job", 0
    start = time.time()
    while time.time() - start < MAX_POLL:
        try:
            r = requests.get(f"{base}/api/v1/call-processing/status/{job_id}", headers=HEADERS, timeout=TIMEOUT)
            d = r.json()
            st = d.get("status","?"); pct = d.get("progress",{}).get("percent","?")
            print(f"    {label} {st} {pct}%")
            if st in ("completed","failed"): return st, round(time.time()-start,1)
        except Exception as e:
            print(f"    {label} poll err: {e}")
        time.sleep(POLL_INTERVAL)
    return "timeout", round(time.time()-start,1)


def fetch_summary(base, call_id):
    try:
        r = requests.get(f"{base}/api/v1/call-processing/summary/{call_id}", headers=HEADERS,
                         params={"include_chunks":"true"}, timeout=TIMEOUT)
        return r.json() if r.status_code == 200 else None
    except: return None


def fetch_detail(base, call_id):
    try:
        r = requests.get(f"{base}/api/v1/call-processing/calls/{call_id}/detail", headers=HEADERS,
                         params={"include_transcript":"true","include_segments":"true"}, timeout=TIMEOUT)
        return r.json() if r.status_code == 200 else None
    except: return None


def extract_rep_name(segments):
    if not segments: return None
    patterns = [r"(?:this is|my name is|i'm|it's|i am)\s+([A-Z][a-z]{2,15})"]
    reps = [s for s in segments[:15] if s.get("speaker") == "customer_rep"]
    for seg in reps[:5]:
        for p in patterns:
            m = re.search(p, seg.get("text",""), re.IGNORECASE)
            if m:
                name = m.group(1).strip().title()
                skip = {"How","What","The","This","That","Well","Yes","Yeah","Okay","Sure","Just","Here","Arizona","Roofers","Thank"}
                if name not in skip and len(name) >= 3: return name
    return None


# ─── Extract row — ALL API fields ───

def extract_row(summary, detail):
    """Extract every field returned by the API into a flat dict."""
    row = {}
    if not summary: return row

    s   = summary.get("summary") or {}
    q   = summary.get("qualification") or {}
    comp_data = summary.get("compliance") or {}
    sop = comp_data.get("sop_compliance") or {}
    obj_data  = summary.get("objections") or {}
    bant  = q.get("bant_scores") or {}
    stages = sop.get("stages") or {}
    addr  = q.get("service_address_structured") or {}

    # ── Detail / transcript ──
    transcript_text = ""
    segments = []
    if detail:
        t = detail.get("transcript", "")
        if isinstance(t, dict):
            segments = t.get("segments", []) or []
            transcript_text = "\n".join(f"{seg.get('speaker','?')}: {seg.get('text','')}" for seg in segments)
        elif isinstance(t, str):
            transcript_text = t
            segments = detail.get("segments", []) or []
        else:
            segments = detail.get("segments", []) or []
            transcript_text = "\n".join(f"{seg.get('speaker','?')}: {seg.get('text','')}" for seg in segments)

    # ── Pending actions ──
    pending = s.get("pending_actions", []) or []
    pa_texts, pa_detail_texts = [], []
    for i, pa in enumerate(pending):
        pa_texts.append(f"[{pa.get('type','')}] {pa.get('action_item','')} (owner: {pa.get('owner','')}, due: {pa.get('due_at','')})")
        pa_detail_texts.append(
            f"PA #{i+1}:\n  Type: {pa.get('type','')}\n  Action: {pa.get('action_item','')}\n"
            f"  Owner: {pa.get('owner','')}\n  Due: {pa.get('due_at','')}\n  Confidence: {pa.get('confidence','')}"
        )

    # Validate pending actions vs transcript
    pa_vs_transcript = "N/A - No actions"
    pa_validation = "No pending actions detected"
    if pending and transcript_text:
        tl = transcript_text.lower()
        all_pass = True
        val_lines = []
        for i, pa in enumerate(pending):
            raw = (pa.get("raw_text") or "").lower()
            action = (pa.get("action_item") or "").lower()
            raw_words = [w for w in raw.split() if len(w) > 3]
            action_words = [w for w in action.split() if len(w) > 3]
            rm = (sum(1 for w in raw_words if w in tl) / max(len(raw_words), 1)) >= 0.5 if raw_words else raw in tl
            am = (sum(1 for w in action_words if w in tl) / max(len(action_words), 1)) >= 0.4 if action_words else False
            ok = rm or am
            if not ok: all_pass = False
            val_lines.append(f"PA #{i+1} [{pa.get('type','')}]: {'PASS' if ok else 'FAIL'}")
        pa_vs_transcript = "PASS" if all_pass else "FAIL"
        pa_validation = "\n".join(val_lines)

    # ── Objections ──
    objs = obj_data.get("objections", []) or []
    obj_texts = []
    for o in objs:
        ov = "Resolved" if o.get("overcome") else "Unresolved"
        obj_texts.append(
            f"[{o.get('category_text','')}] \"{o.get('objection_text','')[:120]}\" "
            f"({ov}, sev={o.get('severity','')}, conf={o.get('confidence_score','')})"
        )
    obj_detail_texts = []
    for i, o in enumerate(objs):
        obj_detail_texts.append(
            f"Obj #{i+1}: [{o.get('category_text','')}]\n"
            f"  Text: {o.get('objection_text','')}\n"
            f"  Sub-objection: {o.get('sub_objection','')}\n"
            f"  Overcome: {o.get('overcome')}\n"
            f"  Severity: {o.get('severity','')}\n"
            f"  Confidence: {o.get('confidence_score','')}\n"
            f"  Timestamp: {o.get('timestamp','')}\n"
            f"  Suggestions: {'; '.join(o.get('response_suggestions') or [])}"
        )

    # ── Tags ──
    tags = set()
    for field in [q.get("qualification_status"), q.get("booking_status"), q.get("detected_call_type"), q.get("service_requested")]:
        if field: tags.add(str(field))

    # ── CSR Agent ──
    rep = extract_rep_name(segments)
    meta = summary.get("metadata") or (detail.get("metadata") if detail else {}) or {}
    agent = (meta.get("agent") or {})
    metadata_rep = agent.get("name") or meta.get("rep_name", "")

    # ═══════════════════════════════════════════════
    # SECTION 1 — CALL OVERVIEW (from detail API)
    # ═══════════════════════════════════════════════
    row["Call_ID"]              = ""  # filled by caller
    row["audio_url"]            = ""  # filled by caller
    row["Call Received"]        = summary.get("processed_at") or (detail.get("call_date") if detail else "")
    row["Processed At"]         = summary.get("processed_at", "")
    row["Created At"]           = detail.get("created_at", "") if detail else ""
    row["Call Date"]            = detail.get("call_date", "") if detail else ""
    row["Status"]               = summary.get("status", "")
    row["Phone number"]         = (detail.get("phone_number", "") if detail else "") or q.get("customer_phone", "")
    row["Rep Role"]             = detail.get("rep_role", "") if detail else ""
    row["Duration (s)"]         = detail.get("duration", "")   if detail else ""
    row["Duration (ms)"]        = detail.get("duration_ms", "") if detail else ""
    row["CSR - Agent"]          = rep or metadata_rep or ""
    row["Agent ID"]             = agent.get("id", "")
    row["Agent Email"]          = agent.get("email", "") or meta.get("rep_email", "")

    # ═══════════════════════════════════════════════
    # SECTION 2 — QUALIFICATION
    # ═══════════════════════════════════════════════
    row["Customer"]                     = q.get("customer_name", "")
    row["Customer Name Confidence"]     = q.get("customer_name_confidence", "")
    row["Customer_type"]                = "Existing" if q.get("is_existing_customer") else "New" if q.get("is_existing_customer") is False else ""
    row["Decision Makers"]              = ", ".join(str(d) for d in (q.get("decision_makers") or []))
    row["Qualified"]                    = q.get("qualification_status", "")
    row["BANT Overall Score"]           = q.get("overall_score", "")
    row["BANT Budget"]                  = bant.get("budget", "")
    row["BANT Authority"]               = bant.get("authority", "")
    row["BANT Need"]                    = bant.get("need", "")
    row["BANT Timeline"]                = bant.get("timeline", "")
    row["Urgency Signals"]              = "\n".join(q.get("urgency_signals") or [])
    row["Budget Indicators"]            = "\n".join(q.get("budget_indicators") or [])
    row["Qualification Confidence"]     = q.get("confidence_score", "")

    # ═══════════════════════════════════════════════
    # SECTION 3 — CALL TYPE & SCOPE
    # ═══════════════════════════════════════════════
    row["Call Type"]                    = q.get("detected_call_type", "")
    row["Scope Classification"]         = q.get("scope_classification", "")
    row["Scope Reason"]                 = q.get("scope_reason", "")
    row["Scope Confidence"]             = q.get("scope_confidence", "")
    row["Scope Signals"]                = "\n".join(q.get("scope_signals") or [])
    row["Service Offered"]              = q.get("service_requested", "")
    row["Service Not Offered Reason"]   = q.get("service_not_offered_reason", "")
    row["Deferred Reason"]              = q.get("deferred_reason", "")
    row["Call Outcome"]                 = q.get("call_outcome_category", "")

    # ═══════════════════════════════════════════════
    # SECTION 4 — BOOKING & APPOINTMENT
    # ═══════════════════════════════════════════════
    row["Booked"]                       = q.get("booking_status", "")
    row["Appointment Confirmed"]        = "Yes" if q.get("appointment_confirmed") else "No"
    row["Appointment Date"]             = q.get("appointment_date", "")
    row["Appointment Type"]             = q.get("appointment_type", "")
    row["Appointment Timezone"]         = q.get("appointment_timezone", "")
    row["Appointment Time Confidence"]  = q.get("appointment_time_confidence", "")
    row["Preferred Time Window"]        = q.get("preferred_time_window", "")
    row["Appointment Intent"]           = q.get("appointment_intent", "")
    row["Original Appointment DateTime"]= q.get("original_appointment_datetime", "")
    row["New Requested Time"]           = q.get("new_requested_time", "")

    # ═══════════════════════════════════════════════
    # SECTION 5 — SERVICE ADDRESS
    # ═══════════════════════════════════════════════
    row["Service Address"]              = q.get("service_address_raw", "")
    row["Address Line1"]                = addr.get("line1", "")
    row["Address City"]                 = addr.get("city", "")
    row["Address State"]                = addr.get("state", "")
    row["Address Postal Code"]          = addr.get("postal_code", "")
    row["Address Country"]              = addr.get("country", "")
    row["Address Confidence"]           = q.get("address_confidence", "")

    # ═══════════════════════════════════════════════
    # SECTION 6 — FOLLOW-UP
    # ═══════════════════════════════════════════════
    row["Follow Up Required"]           = q.get("follow_up_required", "")
    row["Follow Up Reason"]             = q.get("follow_up_reason", "")

    # ═══════════════════════════════════════════════
    # SECTION 7 — SUMMARY
    # ═══════════════════════════════════════════════
    row["Summary"]                      = s.get("summary", "")
    row["Key Points"]                   = "\n".join(s.get("key_points", []) or [])
    row["Action Items"]                 = "\n".join(s.get("action_items", []) or [])
    row["Next Steps"]                   = "\n".join(s.get("next_steps", []) or [])
    row["Sentiment Score"]              = s.get("sentiment_score", "")
    row["Summary Confidence Score"]     = s.get("confidence_score", "")

    # ═══════════════════════════════════════════════
    # SECTION 8 — COMPLIANCE / SOP
    # ═══════════════════════════════════════════════
    row["Compliance Score"]             = sop.get("score", "")
    row["Compliance Rate"]              = sop.get("compliance_rate", "")
    row["Compliance Confidence"]        = sop.get("confidence", "")
    row["Compliance Target Role"]       = comp_data.get("target_role", "")
    row["Compliance Eval Mode"]         = comp_data.get("evaluation_mode", "")
    row["Stages Total"]                 = (stages.get("total", "") if isinstance(stages, dict) else "")
    row["Stages Followed"]              = ", ".join(stages.get("followed", []) or []) if isinstance(stages, dict) else ""
    row["Stages Missed"]                = ", ".join(stages.get("missed", []) or []) if isinstance(stages, dict) else ""
    row["Compliance Issues"]            = "\n".join(sop.get("issues") or [])
    row["Positive Behaviors"]           = "\n".join(sop.get("positive_behaviors") or [])
    row["Coaching Issues"]              = "\n".join(
        f"[{ci.get('severity','')}] {ci.get('issue','')}" for ci in (sop.get("coaching_issues") or [])
    )
    row["Coaching Strengths"]           = "\n".join(
        f"{cs.get('behavior','')}" for cs in (sop.get("coaching_strengths") or [])
    )

    # ═══════════════════════════════════════════════
    # SECTION 9 — OBJECTIONS
    # ═══════════════════════════════════════════════
    row["Objections Count"]             = obj_data.get("total_count", len(objs))
    row["Objections and references"]    = "\n".join(obj_texts) if obj_texts else "No"
    row["Objections Detail"]            = "\n\n".join(obj_detail_texts)

    # ═══════════════════════════════════════════════
    # SECTION 10 — PENDING ACTIONS
    # ═══════════════════════════════════════════════
    row["Pending Actions Count"]        = len(pending)
    row["Pending Actions"]              = "\n".join(pa_texts)
    row["Pending Actions Detail"]       = "\n\n".join(pa_detail_texts)
    row["Pending Actions vs Transcript"]= pa_vs_transcript
    row["Pending Actions Validation"]   = pa_validation

    # ═══════════════════════════════════════════════
    # SECTION 11 — TRANSCRIPT & TAGS
    # ═══════════════════════════════════════════════
    row["Transcript"]                   = transcript_text
    row["Tags"]                         = ", ".join(tags)

    return row


# ─── Excel Writer ───

THIN = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
WRAP = Alignment(vertical="top", wrap_text=True)
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
H_FONT = Font(bold=True, color="FFFFFF", size=10)

BLUE   = PatternFill("solid", fgColor="1F4E79")
GREEN  = PatternFill("solid", fgColor="548235")
GOLD   = PatternFill("solid", fgColor="BF8F00")
RED    = PatternFill("solid", fgColor="C00000")
PASS_F = PatternFill("solid", fgColor="C6EFCE")
WARN_F = PatternFill("solid", fgColor="FFEB9C")
FAIL_F = PatternFill("solid", fgColor="FFC7CE")
MATCH_F = PatternFill("solid", fgColor="E2EFDA")
MISMATCH_F = PatternFill("solid", fgColor="FCE4EC")

# All API fields — grouped by section
REF_COLS = [
    # CALL OVERVIEW
    ("Call_ID", 32), ("audio_url", 38), ("Call Received", 20), ("Processed At", 20),
    ("Created At", 20), ("Call Date", 18), ("Status", 12), ("Phone number", 16),
    ("Rep Role", 14), ("Duration (s)", 10), ("Duration (ms)", 12),
    ("CSR - Agent", 16), ("Agent ID", 28), ("Agent Email", 28),
    # QUALIFICATION
    ("Customer", 22), ("Customer Name Confidence", 12), ("Customer_type", 12),
    ("Decision Makers", 20), ("Qualified", 14), ("BANT Overall Score", 10),
    ("BANT Budget", 8), ("BANT Authority", 8), ("BANT Need", 8), ("BANT Timeline", 8),
    ("Urgency Signals", 35), ("Budget Indicators", 35), ("Qualification Confidence", 10),
    # CALL TYPE & SCOPE
    ("Call Type", 16), ("Scope Classification", 16), ("Scope Reason", 40),
    ("Scope Confidence", 10), ("Scope Signals", 35), ("Service Offered", 30),
    ("Service Not Offered Reason", 30), ("Deferred Reason", 25), ("Call Outcome", 20),
    # BOOKING & APPOINTMENT
    ("Booked", 14), ("Appointment Confirmed", 12), ("Appointment Date", 20),
    ("Appointment Type", 16), ("Appointment Timezone", 14), ("Appointment Time Confidence", 10),
    ("Preferred Time Window", 20), ("Appointment Intent", 14),
    ("Original Appointment DateTime", 22), ("New Requested Time", 22),
    # SERVICE ADDRESS
    ("Service Address", 35), ("Address Line1", 25), ("Address City", 16),
    ("Address State", 10), ("Address Postal Code", 12), ("Address Country", 10),
    ("Address Confidence", 10),
    # FOLLOW-UP
    ("Follow Up Required", 10), ("Follow Up Reason", 40),
    # SUMMARY
    ("Summary", 55), ("Key Points", 45), ("Action Items", 35), ("Next Steps", 35),
    ("Sentiment Score", 10), ("Summary Confidence Score", 10),
    # COMPLIANCE / SOP
    ("Compliance Score", 10), ("Compliance Rate", 10), ("Compliance Confidence", 10),
    ("Compliance Target Role", 16), ("Compliance Eval Mode", 14),
    ("Stages Total", 8), ("Stages Followed", 40), ("Stages Missed", 40),
    ("Compliance Issues", 50), ("Positive Behaviors", 50),
    ("Coaching Issues", 50), ("Coaching Strengths", 50),
    # OBJECTIONS
    ("Objections Count", 10), ("Objections and references", 50), ("Objections Detail", 60),
    # PENDING ACTIONS
    ("Pending Actions Count", 10), ("Pending Actions", 45), ("Pending Actions Detail", 55),
    ("Pending Actions vs Transcript", 12), ("Pending Actions Validation", 40),
    # TRANSCRIPT & TAGS
    ("Transcript", 70), ("Tags", 30),
]

# Comparison tab columns
CMP_COLS = [
    ("Call #", 6), ("audio_url", 30),
    ("PROD Call_ID", 28), ("STAGING Call_ID", 28), ("Verdict", 10), ("Failure Reason", 55),
    ("PROD Summary", 50), ("STAGING Summary", 50),
    ("PROD Customer", 18), ("STAGING Customer", 18),
    ("PROD Booked", 12), ("STAGING Booked", 12), ("Booking Match", 12),
    ("PROD Call Type", 14), ("STAGING Call Type", 14), ("Call Type Match", 12),
    ("PROD Qualified", 12), ("STAGING Qualified", 12),
    ("PROD Call Outcome", 18), ("STAGING Call Outcome", 18),
    ("PROD Compliance", 10), ("STAGING Compliance", 10), ("Compliance Delta", 10),
    ("PROD Sentiment", 10), ("STAGING Sentiment", 10), ("Sentiment Delta", 10),
    ("PROD BANT Budget", 8), ("STAGING BANT Budget", 8),
    ("PROD BANT Authority", 8), ("STAGING BANT Authority", 8),
    ("PROD BANT Need", 8), ("STAGING BANT Need", 8),
    ("PROD BANT Timeline", 8), ("STAGING BANT Timeline", 8), ("BANT Timeline Delta", 10),
    ("PROD BANT Overall", 8), ("STAGING BANT Overall", 8), ("BANT Overall Delta", 10),
    ("PROD Lead Score", 8), ("STAGING Lead Score", 8),
    ("PROD Objections Count", 10), ("STAGING Objections Count", 10),
    ("PROD Scope", 14), ("STAGING Scope", 14),
    ("PROD Appt Confirmed", 10), ("STAGING Appt Confirmed", 10),
]


SECTION_GROUPS = [
    ("CALL OVERVIEW", ["Call_ID","audio_url","Call Received","Processed At","Created At","Call Date","Status","Phone number","Rep Role","Duration (s)","Duration (ms)","CSR - Agent","Agent ID","Agent Email"]),
    ("QUALIFICATION", ["Customer","Customer Name Confidence","Customer_type","Decision Makers","Qualified","BANT Overall Score","BANT Budget","BANT Authority","BANT Need","BANT Timeline","Urgency Signals","Budget Indicators","Qualification Confidence"]),
    ("CALL TYPE & SCOPE", ["Call Type","Scope Classification","Scope Reason","Scope Confidence","Scope Signals","Service Offered","Service Not Offered Reason","Deferred Reason","Call Outcome"]),
    ("BOOKING & APPOINTMENT", ["Booked","Appointment Confirmed","Appointment Date","Appointment Type","Appointment Timezone","Appointment Time Confidence","Preferred Time Window","Appointment Intent","Original Appointment DateTime","New Requested Time"]),
    ("SERVICE ADDRESS", ["Service Address","Address Line1","Address City","Address State","Address Postal Code","Address Country","Address Confidence"]),
    ("FOLLOW-UP", ["Follow Up Required","Follow Up Reason"]),
    ("SUMMARY", ["Summary","Key Points","Action Items","Next Steps","Sentiment Score","Summary Confidence Score"]),
    ("COMPLIANCE / SOP", ["Compliance Score","Compliance Rate","Compliance Confidence","Compliance Target Role","Compliance Eval Mode","Stages Total","Stages Followed","Stages Missed","Compliance Issues","Positive Behaviors","Coaching Issues","Coaching Strengths"]),
    ("OBJECTIONS", ["Objections Count","Objections and references","Objections Detail"]),
    ("PENDING ACTIONS", ["Pending Actions Count","Pending Actions","Pending Actions Detail","Pending Actions vs Transcript","Pending Actions Validation"]),
    ("TRANSCRIPT & TAGS", ["Transcript","Tags"]),
]

SECTION_FILLS = {
    "CALL OVERVIEW":        PatternFill("solid", fgColor="1F4E79"),
    "QUALIFICATION":        PatternFill("solid", fgColor="375623"),
    "CALL TYPE & SCOPE":    PatternFill("solid", fgColor="7B3F00"),
    "BOOKING & APPOINTMENT":PatternFill("solid", fgColor="4A235A"),
    "SERVICE ADDRESS":      PatternFill("solid", fgColor="1A5276"),
    "FOLLOW-UP":            PatternFill("solid", fgColor="922B21"),
    "SUMMARY":              PatternFill("solid", fgColor="0B5345"),
    "COMPLIANCE / SOP":     PatternFill("solid", fgColor="4D4D4D"),
    "OBJECTIONS":           PatternFill("solid", fgColor="784212"),
    "PENDING ACTIONS":      PatternFill("solid", fgColor="154360"),
    "TRANSCRIPT & TAGS":    PatternFill("solid", fgColor="212F3C"),
}


def write_data_sheet(wb, title, rows, header_fill):
    ws = wb.create_sheet(title)
    ws.freeze_panes = "C3"  # freeze past call_id + audio_url, past header rows

    col_to_section = {}
    for sec, fields in SECTION_GROUPS:
        for f in fields:
            col_to_section[f] = sec

    # Row 1: section labels (merged spans)
    # Row 2: field names
    col_idx = 1
    for sec, fields in SECTION_GROUPS:
        sec_fill = SECTION_FILLS.get(sec, BLUE)
        start_col = col_idx
        for field in fields:
            ws.column_dimensions[get_column_letter(col_idx)].width = dict(REF_COLS).get(field, 18)
            # Row 2 — field header
            c2 = ws.cell(row=2, column=col_idx, value=field)
            c2.font = H_FONT; c2.fill = sec_fill; c2.alignment = CENTER_WRAP; c2.border = THIN
            col_idx += 1
        # Row 1 — section header (merged)
        end_col = col_idx - 1
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        c1 = ws.cell(row=1, column=start_col, value=sec)
        c1.font = Font(bold=True, color="FFFFFF", size=11)
        c1.fill = sec_fill
        c1.alignment = Alignment(horizontal="center", vertical="center")
        c1.border = THIN

    # Data rows start at row 3
    col_names = [name for name, _ in REF_COLS]
    for ri, row in enumerate(rows, 3):
        for ci, name in enumerate(col_names, 1):
            val = row.get(name, "")
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = WRAP; c.border = THIN
        # Alternate row shading
        if ri % 2 == 0:
            for ci in range(1, len(col_names) + 1):
                ws.cell(row=ri, column=ci).fill = PatternFill("solid", fgColor="F5F5F5")

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 30
    return ws


def safe_delta(v1, v2):
    if v1 is not None and v2 is not None and v1 != "" and v2 != "":
        try: return round(abs(float(v2) - float(v1)), 3)
        except: pass
    return ""

def safe_match(v1, v2):
    a, b = str(v1 or "").strip().lower(), str(v2 or "").strip().lower()
    if not a and not b: return "N/A"
    return "MATCH" if a == b else "MISMATCH"


def write_comparison_sheet(wb, prod_rows, staging_rows):
    ws = wb.create_sheet("Comparison")
    ws.freeze_panes = "A2"
    for ci, (name, width) in enumerate(CMP_COLS, 1):
        c = ws.cell(row=1, column=ci, value=name)
        c.font = H_FONT; c.fill = RED; c.alignment = CENTER_WRAP; c.border = THIN
        ws.column_dimensions[get_column_letter(ci)].width = width

    for ri, (pr, sr) in enumerate(zip(prod_rows, staging_rows), 2):
        mb = safe_match(pr.get("Booked"), sr.get("Booked"))
        mc = safe_match(pr.get("Call Type"), sr.get("Call Type"))
        dc = safe_delta(pr.get("Compliance Score"), sr.get("Compliance Score"))
        ds = safe_delta(pr.get("Sentiment Score"), sr.get("Sentiment Score"))
        dbt = safe_delta(pr.get("BANT Timeline"), sr.get("BANT Timeline"))
        dbo = safe_delta(pr.get("BANT Overall"), sr.get("BANT Overall"))

        # Verdict
        issues = []
        if mb == "MISMATCH": issues.append(f"Booking: {pr.get('Booked')} vs {sr.get('Booked')}")
        if mc == "MISMATCH": issues.append(f"Call Type: {pr.get('Call Type')} vs {sr.get('Call Type')}")
        if isinstance(dc, float) and dc > 0.15: issues.append(f"Compliance delta={dc}")
        if isinstance(dbt, float) and dbt > 0.2: issues.append(f"BANT Timeline delta={dbt}")
        if isinstance(dbo, float) and dbo > 0.2: issues.append(f"BANT Overall delta={dbo}")
        try:
            oc = int(pr.get("Objections Count") or 0); gc = int(sr.get("Objections Count") or 0)
            if abs(gc - oc) > 2: issues.append(f"Objections: {oc} vs {gc}")
        except: pass

        # Check if staging has actual data (not just a call_id with empty fields)
        staging_has_data = bool(sr.get("Booked") or sr.get("Call Type") or sr.get("Qualified") or sr.get("Summary"))
        prod_has_data    = bool(pr.get("Booked") or pr.get("Call Type") or pr.get("Qualified") or pr.get("Summary"))

        has_high = any("Booking" in i or "Call Type" in i for i in issues)
        if not pr.get("Call_ID") or not sr.get("Call_ID"):
            verdict = "N/A"; failure = "Missing data"
        elif not staging_has_data:
            verdict = "N/A"; failure = "Staging API returned no data (processing error)"
        elif not prod_has_data:
            verdict = "N/A"; failure = "Production API returned no data (processing error)"
        elif has_high:
            verdict = "FAIL"; failure = " | ".join(issues)
        elif len(issues) > 2:
            verdict = "WARN"; failure = " | ".join(issues)
        elif issues:
            verdict = "PASS*"; failure = " | ".join(issues)
        else:
            verdict = "PASS"; failure = ""

        vals = [
            ri - 1, pr.get("audio_url", sr.get("audio_url", "")),
            pr.get("Call_ID", ""), sr.get("Call_ID", ""), verdict, failure,
            str(pr.get("Summary", ""))[:300], str(sr.get("Summary", ""))[:300],
            pr.get("Customer", ""), sr.get("Customer", ""),
            pr.get("Booked", ""), sr.get("Booked", ""), mb,
            pr.get("Call Type", ""), sr.get("Call Type", ""), mc,
            pr.get("Qualified", ""), sr.get("Qualified", ""),
            pr.get("Call Outcome", ""), sr.get("Call Outcome", ""),
            pr.get("Compliance Score"), sr.get("Compliance Score"), dc,
            pr.get("Sentiment Score"), sr.get("Sentiment Score"), ds,
            pr.get("BANT Budget"), sr.get("BANT Budget"),
            pr.get("BANT Authority"), sr.get("BANT Authority"),
            pr.get("BANT Need"), sr.get("BANT Need"),
            pr.get("BANT Timeline"), sr.get("BANT Timeline"), dbt,
            pr.get("BANT Overall"), sr.get("BANT Overall"), dbo,
            pr.get("Lead Score"), sr.get("Lead Score"),
            pr.get("Objections Count"), sr.get("Objections Count"),
            pr.get("Scope Classification"), sr.get("Scope Classification"),
            pr.get("Appointment Confirmed"), sr.get("Appointment Confirmed"),
        ]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = WRAP; c.border = THIN

        # Color verdict
        vc = ws.cell(row=ri, column=5)
        if verdict == "FAIL": vc.fill = FAIL_F; vc.font = Font(bold=True, color="9C0006")
        elif verdict == "WARN": vc.fill = WARN_F; vc.font = Font(bold=True, color="9C6500")
        elif verdict.startswith("PASS"): vc.fill = PASS_F; vc.font = Font(bold=True, color="006100")
        if failure: ws.cell(row=ri, column=6).fill = FAIL_F

        # Color match columns
        for col_idx in [13, 16]:  # Booking Match, Call Type Match
            mc_cell = ws.cell(row=ri, column=col_idx)
            if mc_cell.value == "MISMATCH": mc_cell.fill = MISMATCH_F; mc_cell.font = Font(bold=True, color="C00000")
            elif mc_cell.value == "MATCH": mc_cell.fill = MATCH_F
    return ws


# ─── Main ───

def main():
    print("=" * 70)
    print("  OTTO — SUBMIT TO BOTH ENVIRONMENTS + COMPARE")
    print(f"  Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  URLs: {len(AUDIO_URLS)}")
    print(f"  Production: {PROD_URL}")
    print(f"  Staging:    {STAGING_URL}")
    print("=" * 70)

    # ── Step 1: Submit to PRODUCTION ──
    print("\n[1/6] Submitting to PRODUCTION")
    prod_jobs = []
    for i, url in enumerate(AUDIO_URLS):
        cid, jid = submit(PROD_URL, url, i, "prod_test")
        prod_jobs.append((url, cid, jid))

    # ── Step 2: Submit to STAGING ──
    print("\n[2/6] Submitting to STAGING")
    staging_jobs = []
    for i, url in enumerate(AUDIO_URLS):
        cid, jid = submit(STAGING_URL, url, i, "stage_test")
        staging_jobs.append((url, cid, jid))

    # ── Step 3: Poll PRODUCTION ──
    print("\n[3/6] Polling PRODUCTION")
    prod_status = {}
    for url, cid, jid in prod_jobs:
        st, dur = poll(PROD_URL, jid, f"[PROD {cid}]")
        prod_status[url] = (cid, st, dur)
        print(f"  {cid}: {st} ({dur}s)")

    # ── Step 4: Poll STAGING ──
    print("\n[4/6] Polling STAGING")
    staging_status = {}
    for url, cid, jid in staging_jobs:
        st, dur = poll(STAGING_URL, jid, f"[STAGE {cid}]")
        staging_status[url] = (cid, st, dur)
        print(f"  {cid}: {st} ({dur}s)")

    # ── Step 5: Fetch all data ──
    print("\n[5/6] Fetching Results")
    prod_rows = []
    staging_rows = []

    for i, url in enumerate(AUDIO_URLS):
        print(f"\n  Call {i+1}/30:")

        # Production
        p_cid, p_st, _ = prod_status.get(url, (None, "not_submitted", 0))
        if p_cid and p_st in ("completed", "timeout"):
            print(f"    PROD: Fetching {p_cid} (status={p_st})...")
            ps = fetch_summary(PROD_URL, p_cid)
            pd = fetch_detail(PROD_URL, p_cid)
            pr = extract_row(ps, pd)
            if not pr and p_st == "timeout":
                pr = {}
                print(f"    PROD: timeout — no data available for {p_cid}")
        else:
            pr = {}
            print(f"    PROD: {p_st}")
        pr["audio_url"] = url
        pr["Call_ID"] = p_cid or ""
        prod_rows.append(pr)

        # Staging
        s_cid, s_st, _ = staging_status.get(url, (None, "not_submitted", 0))
        if s_cid and s_st in ("completed", "timeout"):
            print(f"    STAGING: Fetching {s_cid} (status={s_st})...")
            ss = fetch_summary(STAGING_URL, s_cid)
            sd = fetch_detail(STAGING_URL, s_cid)
            sr = extract_row(ss, sd)
            if not sr and s_st == "timeout":
                sr = {}
                print(f"    STAGING: timeout — no data available for {s_cid}")
        else:
            sr = {}
            print(f"    STAGING: {s_st}")
        sr["audio_url"] = url
        sr["Call_ID"] = s_cid or ""
        staging_rows.append(sr)

    # ── Step 6: Generate Excel ──
    print("\n[6/6] Generating Reports")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    write_data_sheet(wb, "Staging (Gemini)", staging_rows, GOLD)
    write_data_sheet(wb, "Production (OpenAI)", prod_rows, GREEN)
    write_comparison_sheet(wb, prod_rows, staging_rows)

    wb.save(EXCEL_OUT)
    print(f"\nExcel saved: {EXCEL_OUT}")

    # Summary
    p_ok = sum(1 for v in prod_status.values() if v[1] == "completed")
    s_ok = sum(1 for v in staging_status.values() if v[1] == "completed")
    print(f"\n{'='*70}")
    print(f"  DONE")
    print(f"  Production completed: {p_ok}/30")
    print(f"  Staging completed:    {s_ok}/30")
    print(f"  Report: {EXCEL_OUT}")
    print(f"{'='*70}")


if __name__ == "__main__":
    main()
