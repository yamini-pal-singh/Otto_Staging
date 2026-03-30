#!/usr/bin/env python3
"""
Staging-only test: Submit 10 URLs to staging, fetch all data, generate Excel
with exact column format matching the reference sheet.
"""
import os, sys, uuid, time, json, re
from datetime import datetime

import requests

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("pip install openpyxl"); sys.exit(1)

# ─── Config ───
STAGING_URL = "https://ottoai-stage.shunyalabs.ai"
API_KEY     = "5q3fwliU9ZFo3epTCsUfUiDw1Dy4DnBP"
COMPANY_ID  = "1be5ea90-d3ae-4b03-8b05-f5679cd73bc4"
HEADERS     = {"X-API-Key": API_KEY, "Content-Type": "application/json"}
TIMEOUT     = 60
POLL_INTERVAL = 15
MAX_POLL     = 1800  # 30 min

OUTPUT_DIR  = os.path.dirname(os.path.abspath(__file__))
EXCEL_OUT   = os.path.join(OUTPUT_DIR, "Staging_Gemini_10_Calls.xlsx")

AUDIO_URLS = [
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/216dcba6-09be-4d79-a38e-feab3ed3b430/4117630448.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/d45bff8a-290b-4a5c-8eb0-42516800387c/4117029440.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/3a7a8e5e-8489-46ee-9d43-f9ba02eae7a3/4116085385.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/f263ee14-32d7-4b01-a137-54deadd9bcd3/4116016640.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/a803c564-048b-4b16-937c-71631ecff231/4116006611.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/cd57b467-1cda-4977-a055-8cbae8d5f2d5/4116008258.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/f1c83d50-795d-4e95-8118-f7bad282270a/4115972252.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/58d5dc4c-911d-4a2a-85a7-b396f8d3efef/4115974295.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/fa63c546-bd86-48b7-a2f7-dbda74eff5f0/4116804701.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/3849971b-360f-4a25-8469-cf3985db737b/4115952143.mp3",
]

# ─── API helpers ───

def submit(audio_url, idx):
    call_id = f"stg10_{idx+1}_{uuid.uuid4().hex[:8]}"
    payload = {
        "call_id": call_id, "company_id": COMPANY_ID, "audio_url": audio_url,
        "phone_number": "+14805551234", "rep_role": "customer_rep", "allow_reprocess": True,
        "metadata": {"agent": {"id": "USR_ANTHONY_ARIZONA", "name": "Anthony", "email": "anthony@arizonaroofers.com"}},
    }
    try:
        r = requests.post(f"{STAGING_URL}/api/v1/call-processing/process", headers=HEADERS, json=payload, timeout=TIMEOUT)
        d = r.json()
        print(f"  [{idx+1:2d}] HTTP {r.status_code} | {call_id} | job={d.get('job_id','N/A')}")
        return call_id, d.get("job_id")
    except Exception as e:
        print(f"  [{idx+1:2d}] ERROR: {e}")
        return call_id, None


def poll(job_id, label=""):
    if not job_id: return "no_job", 0
    start = time.time()
    while time.time() - start < MAX_POLL:
        try:
            r = requests.get(f"{STAGING_URL}/api/v1/call-processing/status/{job_id}", headers=HEADERS, timeout=TIMEOUT)
            d = r.json()
            st = d.get("status", "?"); pct = d.get("progress", {}).get("percent", "?")
            print(f"    {label} {st} {pct}%")
            if st in ("completed", "failed"): return st, round(time.time() - start, 1)
        except Exception as e:
            print(f"    {label} poll err: {e}")
        time.sleep(POLL_INTERVAL)
    return "timeout", round(time.time() - start, 1)


def fetch_summary(call_id):
    try:
        r = requests.get(f"{STAGING_URL}/api/v1/call-processing/summary/{call_id}", headers=HEADERS,
                         params={"include_chunks": "true"}, timeout=TIMEOUT)
        return r.json() if r.status_code == 200 else None
    except: return None


def fetch_detail(call_id):
    try:
        r = requests.get(f"{STAGING_URL}/api/v1/call-processing/calls/{call_id}/detail", headers=HEADERS,
                         params={"include_transcript": "true", "include_segments": "true"}, timeout=TIMEOUT)
        return r.json() if r.status_code == 200 else None
    except: return None


def extract_rep_name(segments):
    if not segments: return None
    patterns = [r"(?:this is|my name is|i'm|it's|i am)\s+([A-Z][a-z]{2,15})"]
    reps = [s for s in segments[:15] if s.get("speaker") == "customer_rep"]
    for seg in reps[:5]:
        for p in patterns:
            m = re.search(p, seg.get("text", ""), re.IGNORECASE)
            if m:
                name = m.group(1).strip().title()
                skip = {"How","What","The","This","That","Well","Yes","Yeah","Okay","Sure","Just","Here","Arizona","Roofers","Thank"}
                if name not in skip and len(name) >= 3: return name
    return None


# ─── Extract row with EXACT column order from reference sheet ───

# Exact columns from user's reference sheet
SHEET_COLS = [
    ("Call Received", 18),
    ("audio_url", 38),
    ("CSR - Agent", 16),
    ("Customer", 22),
    ("Phone number", 16),
    ("Qualified", 14),
    ("Service Offered", 30),
    ("Booked", 14),
    ("Customer Intelligence & Lead Qualification", 45),
    ("Objections and references", 50),
    ("Pending Actions Count", 12),
    ("Pending Actions", 45),
    ("Pending Actions Detail", 55),
    ("Pending Actions vs Transcript", 14),
    ("Pending Actions Validation", 40),
    ("Tags", 30),
    ("Call Type", 16),
    ("Customer_type", 14),
    ("Old data displayed on dashboard", 30),
    ("Transcript", 70),
    ("Point raised by Tushar on mail", 30),
    ("QC Status", 12),
    ("Reason", 30),
    ("Call_ID", 32),
    ("Comments", 30),
    ("Action Items", 35),
    ("Address City", 16),
    ("Address Confidence", 12),
    ("Address Country", 10),
    ("Address Line1", 25),
    ("Address Postal Code", 12),
    ("Address State", 10),
    ("Appointment Confirmed", 14),
    ("Appointment Date", 20),
    ("Appointment Intent", 14),
    ("Appointment Time Confidence", 12),
    ("Appointment Timezone", 14),
    ("Appointment Type", 16),
    ("BANT - Authority", 10),
    ("BANT - Budget", 10),
    ("BANT - Need", 10),
    ("BANT - Timeline", 10),
    ("BANT Overall Score", 10),
    ("Budget Indicators", 35),
    ("Call Outcome Category", 20),
    ("Coaching Issues", 50),
    ("Coaching Strengths", 50),
    ("Compliance Issues", 50),
    ("Compliance Rate", 12),
    ("Compliance Score", 12),
    ("Confidence Score (Qual)", 12),
    ("Confidence Score (Summary)", 12),
    ("Customer Email", 25),
    ("Customer Name Confidence", 12),
    ("Decision Makers", 20),
    ("Deferred Reason", 25),
    ("Evaluation Mode", 16),
    ("Follow Up Reason", 40),
    ("Follow Up Required", 14),
    ("Key Points", 45),
    ("Next Steps", 35),
    ("Objections Count", 12),
    ("Positive Behaviors", 50),
    ("Preferred Time Window", 20),
    ("Processing Status", 14),
    ("SOP Version", 14),
    ("Scope Classification", 16),
    ("Scope Confidence", 12),
    ("Scope Reason", 40),
    ("Scope Signals", 35),
    ("Sentiment Score", 12),
    ("Service Address (Raw)", 35),
    ("Service Not Offered Reason", 30),
    ("Stages Followed", 40),
    ("Stages Missed", 40),
    ("Stages Total", 10),
    ("Summary", 55),
    ("Target Role", 16),
    ("Urgency Signals", 35),
    ("Test Date", 18),
    ("Call #", 8),
]


def extract_row(summary, detail, audio_url, call_id, call_num):
    """Extract all API fields mapped to the exact reference sheet columns."""
    row = {col: "" for col, _ in SHEET_COLS}

    if not summary:
        row["audio_url"] = audio_url
        row["Call_ID"] = call_id
        row["Call #"] = call_num
        row["Test Date"] = datetime.now().strftime("%Y-%m-%d %H:%M")
        row["Processing Status"] = "failed"
        return row

    s   = summary.get("summary") or {}
    q   = summary.get("qualification") or {}
    comp_data = summary.get("compliance") or {}
    sop = comp_data.get("sop_compliance") or {}
    obj_data  = summary.get("objections") or {}
    bant  = q.get("bant_scores") or {}
    stages = sop.get("stages") or {}
    addr  = q.get("service_address_structured") or {}

    # Transcript
    transcript_text = ""
    segments = []
    if detail:
        t = detail.get("transcript", "")
        if isinstance(t, dict):
            segments = t.get("segments", []) or []
            transcript_text = "\n".join(f"{seg.get('speaker','?')}: {seg.get('text','')}" for seg in segments)
        elif isinstance(t, str):
            transcript_text = t
            segments = detail.get("segments", []) or []
        else:
            segments = detail.get("segments", []) or []
            transcript_text = "\n".join(f"{seg.get('speaker','?')}: {seg.get('text','')}" for seg in segments)

    # Pending actions
    pending = s.get("pending_actions", []) or []
    pa_texts, pa_detail_texts = [], []
    for i, pa in enumerate(pending):
        pa_texts.append(f"[{pa.get('type','')}] {pa.get('action_item','')} (owner: {pa.get('owner','')}, due: {pa.get('due_at','')})")
        pa_detail_texts.append(
            f"PA #{i+1}: Type={pa.get('type','')}, Action={pa.get('action_item','')}, "
            f"Owner={pa.get('owner','')}, Due={pa.get('due_at','')}, Confidence={pa.get('confidence','')}"
        )

    # Validate pending actions vs transcript
    pa_vs_transcript = "N/A - No actions"
    pa_validation = "No pending actions detected"
    if pending and transcript_text:
        tl = transcript_text.lower()
        all_pass = True
        val_lines = []
        for i, pa in enumerate(pending):
            raw = (pa.get("raw_text") or "").lower()
            action = (pa.get("action_item") or "").lower()
            raw_words = [w for w in raw.split() if len(w) > 3]
            action_words = [w for w in action.split() if len(w) > 3]
            rm = (sum(1 for w in raw_words if w in tl) / max(len(raw_words), 1)) >= 0.5 if raw_words else raw in tl
            am = (sum(1 for w in action_words if w in tl) / max(len(action_words), 1)) >= 0.4 if action_words else False
            ok = rm or am
            if not ok: all_pass = False
            val_lines.append(f"PA #{i+1} [{pa.get('type','')}]: {'PASS' if ok else 'FAIL'}")
        pa_vs_transcript = "PASS" if all_pass else "FAIL"
        pa_validation = "\n".join(val_lines)

    # Objections
    objs = obj_data.get("objections", []) or []
    obj_texts = []
    for o in objs:
        ov = "Resolved" if o.get("overcome") else "Unresolved"
        obj_texts.append(
            f"[{o.get('category_text','')}] \"{o.get('objection_text','')[:120]}\" "
            f"({ov}, sev={o.get('severity','')}, conf={o.get('confidence_score','')})"
        )

    # Tags
    tags = set()
    for field in [q.get("qualification_status"), q.get("booking_status"), q.get("detected_call_type"), q.get("service_requested")]:
        if field: tags.add(str(field))

    # CSR Agent
    rep = extract_rep_name(segments)
    meta = summary.get("metadata") or (detail.get("metadata") if detail else {}) or {}
    agent_meta = (meta.get("agent") or {})
    metadata_rep = agent_meta.get("name") or meta.get("rep_name", "")

    # Customer Intelligence & Lead Qualification (combined summary)
    lead_qual_parts = []
    if q.get("qualification_status"): lead_qual_parts.append(f"Status: {q['qualification_status']}")
    if q.get("overall_score") is not None: lead_qual_parts.append(f"BANT Overall: {q['overall_score']}")
    if bant: lead_qual_parts.append(f"Budget={bant.get('budget','')}, Auth={bant.get('authority','')}, Need={bant.get('need','')}, Timeline={bant.get('timeline','')}")
    if q.get("urgency_signals"): lead_qual_parts.append(f"Urgency: {', '.join(q['urgency_signals'])}")
    if q.get("budget_indicators"): lead_qual_parts.append(f"Budget indicators: {', '.join(q['budget_indicators'])}")
    lead_qual = "\n".join(lead_qual_parts)

    # ── Map to exact sheet columns ──
    row["Call Received"]                    = summary.get("processed_at") or (detail.get("call_date") if detail else "")
    row["audio_url"]                        = audio_url
    row["CSR - Agent"]                      = rep or metadata_rep or ""
    row["Customer"]                         = q.get("customer_name", "")
    row["Phone number"]                     = (detail.get("phone_number", "") if detail else "") or ""
    row["Qualified"]                        = q.get("qualification_status", "")
    row["Service Offered"]                  = q.get("service_requested", "")
    row["Booked"]                           = q.get("booking_status", "")
    row["Customer Intelligence & Lead Qualification"] = lead_qual
    row["Objections and references"]        = "\n".join(obj_texts) if obj_texts else "None"
    row["Pending Actions Count"]            = len(pending)
    row["Pending Actions"]                  = "\n".join(pa_texts)
    row["Pending Actions Detail"]           = "\n".join(pa_detail_texts)
    row["Pending Actions vs Transcript"]    = pa_vs_transcript
    row["Pending Actions Validation"]       = pa_validation
    row["Tags"]                             = ", ".join(tags)
    row["Call Type"]                        = q.get("detected_call_type", "")
    row["Customer_type"]                    = "Existing" if q.get("is_existing_customer") else "New" if q.get("is_existing_customer") is False else ""
    row["Old data displayed on dashboard"]  = ""  # manual field
    row["Transcript"]                       = transcript_text
    row["Point raised by Tushar on mail"]   = ""  # manual field
    row["QC Status"]                        = ""  # manual field
    row["Reason"]                           = ""  # manual field
    row["Call_ID"]                          = call_id
    row["Comments"]                         = ""  # manual field
    row["Action Items"]                     = "\n".join(s.get("action_items", []) or [])
    row["Address City"]                     = addr.get("city", "")
    row["Address Confidence"]               = q.get("address_confidence", "")
    row["Address Country"]                  = addr.get("country", "")
    row["Address Line1"]                    = addr.get("line1", "")
    row["Address Postal Code"]              = addr.get("postal_code", "")
    row["Address State"]                    = addr.get("state", "")
    row["Appointment Confirmed"]            = "Yes" if q.get("appointment_confirmed") else "No"
    row["Appointment Date"]                 = q.get("appointment_date", "")
    row["Appointment Intent"]               = q.get("appointment_intent", "")
    row["Appointment Time Confidence"]      = q.get("appointment_time_confidence", "")
    row["Appointment Timezone"]             = q.get("appointment_timezone", "")
    row["Appointment Type"]                 = q.get("appointment_type", "")
    row["BANT - Authority"]                 = bant.get("authority", "")
    row["BANT - Budget"]                    = bant.get("budget", "")
    row["BANT - Need"]                      = bant.get("need", "")
    row["BANT - Timeline"]                  = bant.get("timeline", "")
    row["BANT Overall Score"]               = q.get("overall_score", "")
    row["Budget Indicators"]                = "\n".join(q.get("budget_indicators") or [])
    row["Call Outcome Category"]            = q.get("call_outcome_category", "")
    row["Coaching Issues"]                  = "\n".join(f"[{ci.get('severity','')}] {ci.get('issue','')}" for ci in (sop.get("coaching_issues") or []))
    row["Coaching Strengths"]               = "\n".join(f"{cs.get('behavior','')}" for cs in (sop.get("coaching_strengths") or []))
    row["Compliance Issues"]                = "\n".join(sop.get("issues") or [])
    row["Compliance Rate"]                  = sop.get("compliance_rate", "")
    row["Compliance Score"]                 = sop.get("score", "")
    row["Confidence Score (Qual)"]          = q.get("confidence_score", "")
    row["Confidence Score (Summary)"]       = s.get("confidence_score", "")
    row["Customer Email"]                   = ""  # not returned by API
    row["Customer Name Confidence"]         = q.get("customer_name_confidence", "")
    row["Decision Makers"]                  = ", ".join(str(d) for d in (q.get("decision_makers") or []))
    row["Deferred Reason"]                  = q.get("deferred_reason", "")
    row["Evaluation Mode"]                  = comp_data.get("evaluation_mode", "")
    row["Follow Up Reason"]                 = q.get("follow_up_reason", "")
    row["Follow Up Required"]               = q.get("follow_up_required", "")
    row["Key Points"]                       = "\n".join(s.get("key_points", []) or [])
    row["Next Steps"]                       = "\n".join(s.get("next_steps", []) or [])
    row["Objections Count"]                 = obj_data.get("total_count", len(objs))
    row["Positive Behaviors"]               = "\n".join(sop.get("positive_behaviors") or [])
    row["Preferred Time Window"]            = q.get("preferred_time_window", "")
    row["Processing Status"]                = summary.get("status", "")
    row["SOP Version"]                      = ""  # not in API response
    row["Scope Classification"]             = q.get("scope_classification", "")
    row["Scope Confidence"]                 = q.get("scope_confidence", "")
    row["Scope Reason"]                     = q.get("scope_reason", "")
    row["Scope Signals"]                    = "\n".join(q.get("scope_signals") or [])
    row["Sentiment Score"]                  = s.get("sentiment_score", "")
    row["Service Address (Raw)"]            = q.get("service_address_raw", "")
    row["Service Not Offered Reason"]       = q.get("service_not_offered_reason", "")
    row["Stages Followed"]                  = ", ".join(stages.get("followed", []) or []) if isinstance(stages, dict) else ""
    row["Stages Missed"]                    = ", ".join(stages.get("missed", []) or []) if isinstance(stages, dict) else ""
    row["Stages Total"]                     = stages.get("total", "") if isinstance(stages, dict) else ""
    row["Summary"]                          = s.get("summary", "")
    row["Target Role"]                      = comp_data.get("target_role", "")
    row["Urgency Signals"]                  = "\n".join(q.get("urgency_signals") or [])
    row["Test Date"]                        = datetime.now().strftime("%Y-%m-%d %H:%M")
    row["Call #"]                           = call_num

    return row


# ─── Excel Writer ───

THIN = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
WRAP = Alignment(vertical="top", wrap_text=True)
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
H_FONT = Font(bold=True, color="FFFFFF", size=10)
GOLD = PatternFill("solid", fgColor="BF8F00")
ALT_ROW = PatternFill("solid", fgColor="FFF8E7")


def write_sheet(wb, title, rows):
    ws = wb.create_sheet(title)
    ws.freeze_panes = "C2"

    # Header row
    for ci, (name, width) in enumerate(SHEET_COLS, 1):
        c = ws.cell(row=1, column=ci, value=name)
        c.font = H_FONT; c.fill = GOLD; c.alignment = CENTER_WRAP; c.border = THIN
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 30

    # Data rows
    for ri, row in enumerate(rows, 2):
        for ci, (name, _) in enumerate(SHEET_COLS, 1):
            val = row.get(name, "")
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = WRAP; c.border = THIN
        if ri % 2 == 0:
            for ci in range(1, len(SHEET_COLS) + 1):
                ws.cell(row=ri, column=ci).fill = ALT_ROW
    return ws


# ─── Main ───

def main():
    print("=" * 70)
    print("  STAGING-ONLY TEST — 10 Audio URLs")
    print(f"  Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  Staging: {STAGING_URL}")
    print("=" * 70)

    # Step 1: Submit all 10
    print("\n[1/3] Submitting to STAGING")
    jobs = []
    for i, url in enumerate(AUDIO_URLS):
        cid, jid = submit(url, i)
        jobs.append((url, cid, jid))

    # Step 2: Poll
    print("\n[2/3] Polling for completion")
    statuses = {}
    for url, cid, jid in jobs:
        st, dur = poll(jid, f"[{cid}]")
        statuses[url] = (cid, st, dur)
        print(f"  {cid}: {st} ({dur}s)")

    # Step 3: Fetch and build rows
    print("\n[3/3] Fetching data and generating Excel")
    rows = []
    for i, url in enumerate(AUDIO_URLS):
        cid, st, _ = statuses.get(url, (None, "not_submitted", 0))
        print(f"  Call {i+1}/10: {cid} ({st})")
        if cid and st in ("completed", "timeout"):
            sm = fetch_summary(cid)
            dt = fetch_detail(cid)
            row = extract_row(sm, dt, url, cid, i + 1)
        else:
            row = extract_row(None, None, url, cid or "", i + 1)
        rows.append(row)

    # Write Excel
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    write_sheet(wb, "Staging (Gemini)", rows)
    wb.save(EXCEL_OUT)
    print(f"\nExcel saved: {EXCEL_OUT}")

    ok = sum(1 for v in statuses.values() if v[1] == "completed")
    print(f"\n{'='*70}")
    print(f"  DONE — {ok}/10 completed")
    print(f"  Report: {EXCEL_OUT}")
    print(f"{'='*70}")


if __name__ == "__main__":
    main()
